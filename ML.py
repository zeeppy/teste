"""
Analisador de Produtos para o Mercado Livre (Versão Aprimorada)
------------------------------------------------------------
Este script analisa produtos do PDF, pesquisa no Mercado Livre, e gera recomendações.
Versão aprimorada com extração de PDF robusta, análise IA avançada e geração de kits inteligente.
"""

import os
import sys
import re
import json
import time
import random
import traceback
import warnings
from datetime import datetime

# Bibliotecas externas
import pandas as pd
import numpy as np
import PyPDF2
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import openai
from dotenv import load_dotenv

# Configurações
warnings.filterwarnings('ignore')
load_dotenv()  # Carrega variáveis do arquivo .env

# Configuração da API da OpenAI
openai.api_key = os.getenv("OPENAI_API_KEY")

# Configurações globais
VERSION = "2.0.0"
DEBUG = False  # Modo de depuração

def debug_print(message):
    """Imprime uma mensagem somente no modo de depuração"""
    if DEBUG:
        print(f"DEBUG: {message}")

def log_error(message, error=None):
    """Registra um erro no console"""
    print(f"❌ ERRO: {message}")
    if error:
        print(f"   Detalhes: {str(error)}")
        if DEBUG:
            traceback.print_exc()

def log_warning(message):
    """Registra um aviso no console"""
    print(f"⚠️ AVISO: {message}")

def log_success(message):
    """Registra uma operação bem-sucedida no console"""
    print(f"✅ {message}")

# ------------------------------------------------------
# EXTRAÇÃO DE PRODUTOS DO PDF
# ------------------------------------------------------

def extract_products_from_pdf(pdf_path):
    """
    Extrai produtos de um arquivo PDF.
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        
    Returns:
        Lista de produtos (dicionários com 'description', 'code', 'price')
    """
    if not os.path.exists(pdf_path):
        log_error(f"Arquivo não encontrado: {pdf_path}")
        return []
    
    print(f"Lendo arquivo PDF: {pdf_path}")
    
    try:
        # Ler o PDF
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Verificar se o PDF possui páginas
            if len(pdf_reader.pages) == 0:
                log_error("O PDF não contém páginas")
                return []
            
            # Extrair texto do PDF
            full_text = ""
            for i, page in enumerate(pdf_reader.pages):
                print(f"Processando página {i+1}/{len(pdf_reader.pages)}...")
                try:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + "\n\n"
                except Exception as e:
                    log_error(f"Erro ao extrair texto da página {i+1}", e)
            
            # Exibir amostra do conteúdo
            if full_text:
                sample = full_text[:500] + ("..." if len(full_text) > 500 else "")
                print("\nAmostra do conteúdo do PDF:")
                print("-" * 50)
                print(sample)
                print("-" * 50)
                
                # Procurar pela tabela de produtos
                products = find_product_table(full_text)
                
                if products:
                    log_success(f"Encontrados {len(products)} produtos no PDF")
                    
                    # Exibir exemplos
                    print("\nExemplos de produtos encontrados:")
                    for i, product in enumerate(products[:5]):
                        desc = product.get('description', '')
                        code = product.get('code', '')
                        if code and desc:
                            print(f"  {i+1}. [{code}] {desc}")
                        elif desc:
                            print(f"  {i+1}. {desc}")
                    
                    return products
                else:
                    log_error("Não foi possível identificar produtos no PDF")
                    return []
            else:
                log_error("Não foi possível extrair texto do PDF")
                return []
    
    except Exception as e:
        log_error("Erro ao processar o PDF", e)
        return []

def enhanced_extract_products_from_pdf(pdf_path):
    """
    Versão melhorada da extração de produtos de um arquivo PDF.
    Utiliza múltiplos métodos e combina os resultados.
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        
    Returns:
        Lista de produtos (dicionários com 'description', 'code', 'price')
    """
    if not os.path.exists(pdf_path):
        log_error(f"Arquivo não encontrado: {pdf_path}")
        return []
    
    print(f"Lendo arquivo PDF: {pdf_path}")
    
    try:
        # Ler o PDF
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Verificar se o PDF possui páginas
            if len(pdf_reader.pages) == 0:
                log_error("O PDF não contém páginas")
                return []
            
            # Extrair texto do PDF
            full_text = ""
            for i, page in enumerate(pdf_reader.pages):
                print(f"Processando página {i+1}/{len(pdf_reader.pages)}...")
                try:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + "\n\n"
                except Exception as e:
                    log_error(f"Erro ao extrair texto da página {i+1}", e)
            
            if not full_text:
                log_error("Não foi possível extrair texto do PDF")
                return []
                
            # Usar múltiplos métodos para extração e combiná-los
            methods = [
                ("Método principal", find_product_table),
                ("Método alternativo", extract_products_alternative),
                ("Método por linhas", extract_products_by_line),
                ("Método por padrões de preço", extract_products_by_price_pattern)
            ]
            
            all_products = []
            
            for method_name, method_func in methods:
                print(f"\nTentando extração com {method_name}...")
                products = method_func(full_text)
                
                if products:
                    print(f"✅ {method_name}: Encontrados {len(products)} produtos")
                    all_products.extend(products)
                else:
                    print(f"❌ {method_name}: Nenhum produto encontrado")
            
            # Remover produtos duplicados (baseado na descrição)
            unique_products = []
            seen_descriptions = set()
            
            for product in all_products:
                desc = product.get('description', '').strip().lower()
                code = product.get('code', '')
                
                if desc and (desc not in seen_descriptions or code):
                    seen_descriptions.add(desc)
                    unique_products.append(product)
            
            if unique_products:
                log_success(f"Total de produtos únicos encontrados: {len(unique_products)}")
                return unique_products
            else:
                log_error("Não foi possível identificar produtos no PDF após tentar múltiplos métodos")
                return []
    
    except Exception as e:
        log_error("Erro ao processar o PDF", e)
        return []

def find_product_table(text):
    """
    Encontra e extrai a tabela de produtos de um texto.
    
    Args:
        text: Texto completo do PDF
        
    Returns:
        Lista de produtos encontrados
    """
    # Separar o texto em linhas
    lines = text.split('\n')
    
    # Buscar pelo cabeçalho da tabela de produtos
    header_indices = []
    header_patterns = [
        r'(?:Img|Image).*(?:Item|Código).*Descrição',
        r'Descrição.*(?:Qtde|Quantidade).*Valor',
        r'Código.*Produto.*Descrição'
    ]
    
    for i, line in enumerate(lines):
        for pattern in header_patterns:
            if re.search(pattern, line, re.IGNORECASE):
                header_indices.append(i)
                debug_print(f"Possível cabeçalho de tabela na linha {i+1}: '{line}'")
    
    # Se não encontrou cabeçalhos, tentar método alternativo
    if not header_indices:
        debug_print("Cabeçalho de tabela não encontrado. Tentando método alternativo...")
        return extract_products_alternative(text)
    
    # Começar a extração a partir do primeiro cabeçalho encontrado
    products = []
    start_index = min(header_indices) + 1  # Começar na linha após o cabeçalho
    
    # Processar as linhas subsequentes
    for i in range(start_index, len(lines)):
        line = lines[i].strip()
        
        # Ignorar linhas vazias ou muito curtas
        if not line or len(line) < 5:
            continue
        
        # Ignorar linhas que parecem ser cabeçalhos de página, rodapés, etc.
        if re.match(r'^Página|^Total|^Subtotal|^Emissão|^Pedido', line):
            continue
        
        # Tentar extrair código e descrição do produto
        product = parse_product_line(line)
        if product and product.get('description'):
            products.append(product)
    
    return products

def parse_product_line(line):
    """
    Analisa uma linha para extrair informações do produto.
    
    Args:
        line: Uma linha de texto
        
    Returns:
        Dicionário com informações do produto ou None
    """
    # Padrão para código de produto seguido por descrição
    product_match = re.search(r'^([A-Z0-9]{2,12}[-]?[A-Z0-9]{1,8})\s+(.*?)(?:\s+\d+\s+[\d.,]+|\s*$)', line)
    
    if product_match:
        code = product_match.group(1).strip()
        description = product_match.group(2).strip()
        
        # Extrair preço, se presente
        price = None
        price_match = re.search(r'(\d+[,.]\d+)', line)
        if price_match:
            try:
                price_str = price_match.group(1)
                price = float(price_str.replace('.', '').replace(',', '.'))
            except:
                pass
        
        # Ignorar descrições muito curtas ou que não parecem ser produtos
        if len(description) >= 5 and not is_header_or_footer(description):
            return {
                'code': code,
                'description': description,
                'price': price
            }
    
    # Se não encontrou no formato acima, tentar outros padrões
    parts = re.split(r'\s{2,}|\t', line)
    if len(parts) >= 2:
        # Primeiro item pode ser um código
        if re.match(r'^[A-Z0-9]{2,12}[-]?[A-Z0-9]{1,8}$', parts[0]):
            code = parts[0]
            description = parts[1] if len(parts) > 1 else ""
            
            if len(description) >= 5 and not is_header_or_footer(description):
                return {
                    'code': code,
                    'description': description,
                    'price': None
                }
    
    # Para linhas que não seguem o padrão esperado
    if len(line) >= 10 and not is_header_or_footer(line):
        return {
            'code': None,
            'description': line,
            'price': None
        }
    
    return None

def is_header_or_footer(text):
    """
    Verifica se um texto parece ser um cabeçalho ou rodapé.
    
    Args:
        text: Texto para verificar
        
    Returns:
        True se for um cabeçalho/rodapé, False caso contrário
    """
    if not text:
        return True
    
    # Termos comuns em cabeçalhos e rodapés
    header_terms = [
        "unid", "valor", "quantidade", "qtde", "total", "subtotal", 
        "página", "page", "item", "código", "pedido", "emissão", 
        "cnpj", "cpf", "telefone", "email", "www", "http", "endereço"
    ]
    
    text_lower = text.lower()
    
    # Verificar se contém termos de cabeçalho e é curto
    if len(text) < 25:
        for term in header_terms:
            if term in text_lower:
                return True
    
    return False

def improved_is_header_or_footer(text):
    """
    Verificação melhorada para identificar cabeçalhos ou rodapés.
    
    Args:
        text: Texto para verificar
        
    Returns:
        True se for um cabeçalho/rodapé, False caso contrário
    """
    if not text:
        return True
    
    # Termos comuns em cabeçalhos e rodapés (ampliado)
    header_terms = [
        "unid", "valor", "quantidade", "qtde", "total", "subtotal", 
        "página", "page", "item", "código", "pedido", "emissão", 
        "cnpj", "cpf", "telefone", "email", "www", "http", "endereço",
        "data", "orçamento", "cotação", "nota fiscal", "nf-e", "nfe",
        "cliente", "fornecedor", "contato", "frete", "entrega", "prazo",
        "preço", "valor unit", "catálogo", "referência", "ref"
    ]
    
    text_lower = text.lower()
    
    # Verificar se contém termos de cabeçalho e é curto
    if len(text) < 25:
        for term in header_terms:
            if term in text_lower:
                return True
    
    # Padrões que frequentemente indicam cabeçalhos/rodapés
    header_patterns = [
        r'^página\s+\d+\s+de\s+\d+$',
        r'^pedido\s+n[º°]?\s*\d+',
        r'^emissão:\s+\d{2}/\d{2}/\d{4}',
        r'^data:',
        r'^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}',
        r'^nome:',
        r'^fone:',
        r'^produto\s+descrição\s+valor',
        r'^cod\.?\s+produto',
        r'^nenhum registro encontrado',
        r'^-+$'  # Linhas com apenas hífens
    ]
    
    for pattern in header_patterns:
        if re.search(pattern, text_lower):
            return True
    
    return False

def extract_products_alternative(text):
    """
    Método alternativo para extrair produtos quando o método principal falha.
    
    Args:
        text: Texto completo do PDF
        
    Returns:
        Lista de produtos encontrados
    """
    debug_print("Usando método alternativo para extração de produtos")
    
    lines = text.split('\n')
    products = []
    
    # Procurar por padrões específicos de produtos em cada linha
    for line in lines:
        line = line.strip()
        
        # Ignorar linhas vazias ou muito curtas
        if not line or len(line) < 10:
            continue
        
        # Procurar padrões específicos de produtos
        if re.match(r'^[A-Z0-9]{2,12}[-]?[A-Z0-9]{1,8}\s+[A-Z0-9]', line):
            product = parse_product_line(line)
            if product:
                products.append(product)
        # Procurar por descrições substantivas
        elif re.search(r'\b(MESA|CADEIRA|ARMÁRIO|SOFÁ|CAMA|ESTANTE|GABINETE|KIT)\b', line, re.IGNORECASE):
            if not is_header_or_footer(line):
                products.append({
                    'code': None,
                    'description': line,
                    'price': None
                })
    
    # Remover duplicatas
    unique_products = []
    seen_descriptions = set()
    
    for product in products:
        desc = product.get('description', '')
        if desc and desc not in seen_descriptions:
            seen_descriptions.add(desc)
            unique_products.append(product)
    
    return unique_products

def extract_products_by_line(text):
    """
    Extrai produtos analisando o texto linha por linha com maior flexibilidade.
    
    Args:
        text: Texto completo do PDF
        
    Returns:
        Lista de produtos encontrados
    """
    lines = text.split('\n')
    products = []
    
    # Padrões para identificar linhas de produto
    product_patterns = [
        # Código de produto seguido por descrição
        r'^([A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10})\s+(.*?)(?:\s+[\d.,]+)?$',
        # Descrição seguida por código entre parênteses
        r'^(.*?)\s+\(([A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10})\)(?:\s+[\d.,]+)?$',
        # Descrição seguida por quantidade e preço
        r'^(.*?)\s+(\d+)\s+(?:UN|PC|KG|MT)?\s+([\d.,]+)$'
    ]
    
    # Ignorar linhas que contêm estes termos (cabeçalhos/rodapés)
    ignore_terms = [
        'página', 'total', 'subtotal', 'descrição do produto', 'código', 
        'qtde', 'página', 'emissão', 'pedido', 'valor', 'preço'
    ]
    
    # Expressão regular para encontrar preços
    price_pattern = r'R?\$?\s*([\d.,]+)(?:\s*(?:UN|PC|un|pc|cada))?$'
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Ignorar linhas vazias, muito curtas, ou com termos de cabeçalho/rodapé
        if not line or len(line) < 5:
            continue
            
        if any(term in line.lower() for term in ignore_terms):
            continue
        
        # Usar os diferentes padrões para identificar produtos
        product = None
        
        # Tentar cada padrão
        for pattern in product_patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                # Padrão 1: Código + Descrição
                if len(match.groups()) >= 2 and re.match(r'^[A-Z0-9]{2,15}', match.group(1)):
                    code = match.group(1).strip()
                    description = match.group(2).strip()
                    product = {'code': code, 'description': description, 'price': None}
                    break
                # Padrão 2: Descrição + Código entre parênteses
                elif len(match.groups()) >= 2 and re.match(r'^[A-Z0-9]{2,15}', match.group(2)):
                    description = match.group(1).strip()
                    code = match.group(2).strip()
                    product = {'code': code, 'description': description, 'price': None}
                    break
                # Padrão 3: Descrição + Quantidade + Preço
                elif len(match.groups()) >= 3:
                    description = match.group(1).strip()
                    price_str = match.group(3).strip()
                    try:
                        price = float(price_str.replace('.', '').replace(',', '.'))
                        product = {'code': None, 'description': description, 'price': price}
                        break
                    except:
                        pass
        
        # Se não identificou com os padrões acima, verificar se é uma descrição válida
        if not product and len(line) > 10:
            # Procurar por um preço no final da linha
            price_match = re.search(price_pattern, line)
            price = None
            
            if price_match:
                try:
                    price_str = price_match.group(1)
                    price = float(price_str.replace('.', '').replace(',', '.'))
                    # Remover o preço da descrição
                    description = line[:price_match.start()].strip()
                except:
                    description = line
            else:
                description = line
            
            # Verificar se a descrição parece ser um produto válido
            if len(description) >= 5 and not is_header_or_footer(description):
                # Procurar por um código de produto no início
                code_match = re.match(r'^([A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10})\s+', description)
                if code_match:
                    code = code_match.group(1)
                    description = description[code_match.end():].strip()
                    product = {'code': code, 'description': description, 'price': price}
                else:
                    product = {'code': None, 'description': description, 'price': price}
        
        # Adicionar o produto se for válido
        if product and product.get('description') and len(product['description']) >= 5:
            products.append(product)
    
    return products

def extract_products_by_price_pattern(text):
    """
    Extrai produtos procurando por padrões de preço.
    
    Args:
        text: Texto completo do PDF
        
    Returns:
        Lista de produtos encontrados
    """
    lines = text.split('\n')
    products = []
    
    # Padrão de preço
    price_patterns = [
        r'R\$\s*([\d.,]+)',  # R$ 123,45
        r'([\d.,]+)\s*reais',  # 123,45 reais
        r'([\d.,]+)(?:\s*(?:UN|PC|un|pc|cada))?$'  # 123,45 no final da linha
    ]
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Ignorar linhas vazias ou muito curtas
        if not line or len(line) < 10:
            continue
        
        for pattern in price_patterns:
            price_match = re.search(pattern, line, re.IGNORECASE)
            if price_match:
                try:
                    price_str = price_match.group(1)
                    price = float(price_str.replace('.', '').replace(',', '.'))
                    
                    # Extrair a descrição (tudo antes do preço)
                    description = line[:price_match.start()].strip()
                    
                    # Procurar por um código de produto no início
                    code = None
                    code_match = re.match(r'^([A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10})\s+', description)
                    if code_match:
                        code = code_match.group(1)
                        description = description[code_match.end():].strip()
                    
                    if description and len(description) >= 5 and not is_header_or_footer(description):
                        products.append({
                            'code': code,
                            'description': description,
                            'price': price
                        })
                except:
                    pass
    
    return products

# ------------------------------------------------------
# INTERAÇÃO COM MERCADO LIVRE
# ------------------------------------------------------

def search_mercado_livre(product_name):
    """
    Pesquisa um produto no Mercado Livre.
    
    Args:
        product_name: Nome do produto para pesquisar
        
    Returns:
        Lista de resultados encontrados
    """
    if not product_name or len(product_name) < 3:
        return []
    
    search_url = f"https://www.mercadolivre.com.br/jm/search?as_word={product_name.replace(' ', '%20')}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        print(f"Pesquisando no Mercado Livre: {product_name}")
        response = requests.get(search_url, headers=headers, timeout=10)
        
        if response.status_code != 200:
            log_warning(f"Resposta HTTP {response.status_code} ao pesquisar '{product_name}'")
            return []
        
        # Parse HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Encontrar itens de produto
        items = soup.find_all('li', class_='ui-search-layout__item')
        results = []
        
        for item in items[:10]:  # Limitar a 10 resultados
            try:
                # Título
                title_element = item.find('h2', class_='ui-search-item__title')
                title = title_element.text if title_element else "Sem título"
                
                # Preço
                price_element = item.find('span', class_='price-tag-fraction')
                price = 0
                if price_element:
                    try:
                        price = float(price_element.text.replace('.', '').replace(',', '.'))
                    except:
                        pass
                
                # Link
                link_element = item.find('a', class_='ui-search-link')
                link = link_element['href'] if link_element and 'href' in link_element.attrs else ""
                
                # Vendas
                sold_element = item.find('span', class_='ui-search-item__sales')
                sold_count = 0
                if sold_element:
                    match = re.search(r'\d+', sold_element.text)
                    if match:
                        try:
                            sold_count = int(match.group())
                        except:
                            pass
                
                results.append({
                    'title': title,
                    'price': price,
                    'link': link,
                    'sold_count': sold_count
                })
            except Exception as e:
                debug_print(f"Erro ao processar item do Mercado Livre: {str(e)}")
        
        if results:
            log_success(f"Encontrados {len(results)} resultados para '{product_name}'")
        else:
            log_warning(f"Nenhum resultado encontrado para '{product_name}'")
        
        return results
        
    except Exception as e:
        log_error(f"Erro ao pesquisar '{product_name}'", e)
        return []

def search_mercado_livre_robust(product_name, max_retries=3, backoff_factor=1.5):
    """
    Versão robusta da pesquisa no Mercado Livre com retentativas e adaptação aos padrões do site.
    
    Args:
        product_name: Nome do produto para pesquisar
        max_retries: Número máximo de tentativas em caso de falha
        backoff_factor: Fator de espera entre tentativas
        
    Returns:
        Lista de resultados encontrados
    """
    if not product_name or len(product_name) < 3:
        return []
    
    # Limpar e preparar a consulta para URL
    query = prepare_search_query(product_name)
    search_url = f"https://www.mercadolivre.com.br/jm/search?as_word={query}"
    
    # Rotacionar diferentes user agents para evitar bloqueios
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
    ]
    
    # Usar diferentes parsers como fallback
    parsers = [
        parse_search_results_standard,
        parse_search_results_alternative,
        parse_search_results_minimal
    ]
    
    results = []
    attempt = 0
    
    while attempt < max_retries and not results:
        try:
            # Escolher um user agent aleatoriamente
            headers = {
                'User-Agent': random.choice(user_agents),
                'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
                'Cache-Control': 'no-cache',
                'Pragma': 'no-cache'
            }
            
            wait_time = backoff_factor ** attempt
            if attempt > 0:
                print(f"  Tentativa {attempt+1}/{max_retries}. Aguardando {wait_time:.1f}s...")
                time.sleep(wait_time)
                
            print(f"Pesquisando no Mercado Livre: {product_name}")
            response = requests.get(search_url, headers=headers, timeout=15)
            
            if response.status_code != 200:
                log_warning(f"Resposta HTTP {response.status_code} ao pesquisar '{product_name}'")
                attempt += 1
                continue
            
            # Tentar cada um dos parsers até encontrar resultados
            for parser_index, parser_func in enumerate(parsers):
                try:
                    parser_results = parser_func(response.text, product_name)
                    if parser_results:
                        log_success(f"Encontrados {len(parser_results)} resultados para '{product_name}' (parser {parser_index+1})")
                        return parser_results
                except Exception as parser_error:
                    debug_print(f"Erro no parser {parser_index+1}: {str(parser_error)}")
            
            # Se chegou aqui, nenhum parser encontrou resultados
            log_warning(f"Nenhum resultado encontrado para '{product_name}' após tentar {len(parsers)} parsers")
            attempt += 1
            
        except requests.exceptions.Timeout:
            log_warning(f"Timeout ao pesquisar '{product_name}'. Tentativa {attempt+1}/{max_retries}")
            attempt += 1
        except requests.exceptions.ConnectionError:
            log_warning(f"Erro de conexão ao pesquisar '{product_name}'. Tentativa {attempt+1}/{max_retries}")
            attempt += 1
        except Exception as e:
            log_error(f"Erro ao pesquisar '{product_name}'", e)
            attempt += 1
    
    # Se chegou aqui, todas as tentativas falharam
    if not results:
        log_warning(f"Todas as {max_retries} tentativas falharam para '{product_name}'")
    
    return results

def prepare_search_query(product_name):
    """
    Prepara a consulta para pesquisa, removendo caracteres problemáticos
    e otimizando a busca.
    
    Args:
        product_name: Nome do produto original
        
    Returns:
        Consulta preparada para pesquisa
    """
    # Remover caracteres especiais e limitar comprimento
    query = product_name.strip()
    
    # Remover códigos de produto muito específicos que podem limitar demais a busca
    if re.match(r'^[A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10}\s+', query):
        code_part = re.match(r'^([A-Z0-9]{2,15}[-/]?[A-Z0-9]{0,10})\s+', query).group(1)
        if len(code_part) > 4:
            # Se o código é longo, remover para melhorar resultados da busca
            query = query[len(code_part):].strip()
    
    # Limitar o tamanho da consulta
    if len(query) > 80:
        words = query.split()
        short_query = []
        current_length = 0
        
        # Pegar palavras até chegar perto de 80 caracteres
        for word in words:
            if current_length + len(word) + 1 <= 80:
                short_query.append(word)
                current_length += len(word) + 1
            else:
                break
        
        query = ' '.join(short_query)
    
    # Remover caracteres problemáticos
    query = re.sub(r'[^\w\s\-.,]', ' ', query)
    
    # Remover palavras muito curtas (artigos, preposições, etc.)
    query_words = [word for word in query.split() if len(word) > 2]
    query = ' '.join(query_words)
    
    # Codificar para URL
    query = query.replace(' ', '%20')
    
    return query

def parse_search_results_standard(html, product_name):
    """
    Parser padrão para resultados de pesquisa do Mercado Livre.
    
    Args:
        html: HTML da página de resultados
        product_name: Nome do produto pesquisado (para debug)
        
    Returns:
        Lista de resultados
    """
    soup = BeautifulSoup(html, 'html.parser')
    
    # Encontrar itens de produto
    items = soup.find_all('li', class_='ui-search-layout__item')
    
    if not items:
        debug_print(f"Nenhum item encontrado com seletor 'ui-search-layout__item' para '{product_name}'")
        return []
    
    results = []
    
    for item in items[:10]:  # Limitar a 10 resultados
        try:
            # Título
            title_element = item.find('h2', class_='ui-search-item__title')
            title = title_element.text if title_element else "Sem título"
            
            # Preço
            price_element = item.find('span', class_='price-tag-fraction')
            price = 0
            if price_element:
                try:
                    price = float(price_element.text.replace('.', '').replace(',', '.'))
                except:
                    pass
            
            # Link
            link_element = item.find('a', class_='ui-search-link')
            link = link_element['href'] if link_element and 'href' in link_element.attrs else ""
            
            # Vendas
            sold_element = item.find('span', class_='ui-search-item__sales')
            sold_count = 0
            if sold_element:
                match = re.search(r'\d+', sold_element.text)
                if match:
                    try:
                        sold_count = int(match.group())
                    except:
                        pass
            
            results.append({
                'title': title,
                'price': price,
                'link': link,
                'sold_count': sold_count
            })
        except Exception as e:
            debug_print(f"Erro ao processar item do Mercado Livre: {str(e)}")
    
    return results

def parse_search_results_alternative(html, product_name):
    """
    Parser alternativo para resultados de pesquisa do Mercado Livre.
    Usa seletores diferentes caso o layout do site tenha mudado.
    
    Args:
        html: HTML da página de resultados
        product_name: Nome do produto pesquisado (para debug)
        
    Returns:
        Lista de resultados
    """
    soup = BeautifulSoup(html, 'html.parser')
    
    # Tentar diferentes seletores para itens de produto
    selectors = [
        'div.ui-search-result',
        'div.andes-card',
        'div.ui-search-result__wrapper',
        'li.ui-search-layout__item'
    ]
    
    items = []
    for selector in selectors:
        items = soup.select(selector)
        if items:
            debug_print(f"Encontrados {len(items)} itens com seletor '{selector}'")
            break
    
    if not items:
        debug_print(f"Nenhum item encontrado com seletores alternativos para '{product_name}'")
        return []
    
    results = []
    
    for item in items[:10]:  # Limitar a 10 resultados
        try:
            # Título (testar diferentes seletores)
            title_selectors = ['h2', '.ui-search-item__title', '.ui-search-item__group__element']
            title = "Sem título"
            for selector in title_selectors:
                title_element = item.select_one(selector)
                if title_element:
                    title = title_element.text.strip()
                    break
            
            # Preço (testar diferentes seletores)
            price = 0
            price_selectors = [
                '.price-tag-fraction', 
                '.andes-money-amount__fraction',
                '.ui-search-price__part'
            ]
            
            for selector in price_selectors:
                price_element = item.select_one(selector)
                if price_element:
                    try:
                        price_text = price_element.text.strip()
                        price = float(price_text.replace('.', '').replace(',', '.'))
                        break
                    except:
                        pass
            
            # Link
            link = ""
            link_selectors = ['a.ui-search-link', 'a.ui-search-result__content', 'a']
            for selector in link_selectors:
                link_element = item.select_one(selector)
                if link_element and 'href' in link_element.attrs:
                    link = link_element['href']
                    break
            
            # Vendas
            sold_count = 0
            sold_selectors = [
                '.ui-search-item__sales', 
                '.ui-search-item__group__element--shipping',
                '.ui-search-item__highlights-label'
            ]
            
            for selector in sold_selectors:
                sold_element = item.select_one(selector)
                if sold_element and 'vendido' in sold_element.text.lower():
                    match = re.search(r'\d+', sold_element.text)
                    if match:
                        try:
                            sold_count = int(match.group())
                            break
                        except:
                            pass
            
            results.append({
                'title': title,
                'price': price,
                'link': link,
                'sold_count': sold_count
            })
        except Exception as e:
            debug_print(f"Erro ao processar item do Mercado Livre (parser alternativo): {str(e)}")
    
    return results

def parse_search_results_minimal(html, product_name):
    """
    Parser minimalista que tenta extrair informações básicas mesmo 
    quando o layout do site muda drasticamente.
    
    Args:
        html: HTML da página de resultados
        product_name: Nome do produto pesquisado (para debug)
        
    Returns:
        Lista de resultados
    """
    soup = BeautifulSoup(html, 'html.parser')
    results = []
    
    # Encontrar todos os links que podem ser produtos
    product_links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        # Links de produto geralmente têm este padrão
        if 'mercadolivre.com.br' in href and ('/p/' in href or '/produto/' in href or 'MLB-' in href):
            product_links.append(a)
    
    debug_print(f"Encontrados {len(product_links)} possíveis links de produto (parser minimal)")
    
    # Para cada link, tentar encontrar um título e um preço nas proximidades
    for link in product_links[:10]:  # Limitar a 10 
        try:
            # Pegar o link
            href = link['href']
            
            # Tentar encontrar um título
            title = "Sem título"
            title_candidates = [link.text]
            
            # Procurar em elementos próximos
            parent = link.parent
            for i in range(3):  # Subir até 3 níveis
                if parent:
                    for tag in parent.find_all(['h2', 'h3', 'span', 'div']):
                        if tag.text and len(tag.text.strip()) > 5:
                            title_candidates.append(tag.text.strip())
                    parent = parent.parent
            
            # Escolher o melhor candidato a título
            title_candidates = [t for t in title_candidates if len(t.strip()) >= 5]
            if title_candidates:
                title = max(title_candidates, key=len)
            
            # Tentar encontrar um preço
            price = 0
            price_pattern = re.compile(r'R\$\s*([\d.,]+)')
            
            # Procurar o padrão de preço na vizinhança do link
            parent = link.parent
            for i in range(3):  # Subir até 3 níveis
                if parent:
                    parent_html = str(parent)
                    price_matches = price_pattern.findall(parent_html)
                    if price_matches:
                        try:
                            price_str = price_matches[0]
                            price = float(price_str.replace('.', '').replace(',', '.'))
                            break
                        except:
                            pass
                    parent = parent.parent
            
            results.append({
                'title': title,
                'price': price,
                'link': href,
                'sold_count': 0  # Difícil extrair este dado no modo minimal
            })
        except Exception as e:
            debug_print(f"Erro ao processar item do Mercado Livre (parser minimal): {str(e)}")
    
    return results

def get_product_details(product_link):
    """
    Obtém detalhes de um produto a partir do seu link.
    
    Args:
        product_link: URL do produto
        
    Returns:
        Dicionário com detalhes do vendedor
    """
    if not product_link:
        return {}
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        response = requests.get(product_link, headers=headers, timeout=10)
        
        if response.status_code != 200:
            debug_print(f"Resposta HTTP {response.status_code} ao acessar detalhes do produto")
            return {}
        
        # Parse HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Nível do vendedor
        seller_level = "Não informado"
        seller_level_element = soup.find('span', class_='ui-seller-info__status-info')
        if seller_level_element:
            seller_level = seller_level_element.text
        
        # Vendas
        sales = 0
        sales_element = soup.find('strong', class_='ui-seller-info__sales-number')
        if sales_element:
            match = re.search(r'\d+', sales_element.text)
            if match:
                try:
                    sales = int(match.group())
                except:
                    pass
        
        # Avaliação
        rating = 0
        rating_element = soup.find('span', class_='ui-seller-info__rating-average')
        if rating_element:
            try:
                rating = float(rating_element.text.replace(',', '.'))
            except:
                pass
        
        return {
            'seller_level': seller_level,
            'sales': sales,
            'rating': rating
        }
        
    except Exception as e:
        debug_print(f"Erro ao obter detalhes do produto: {str(e)}")
        return {}

def get_product_details_robust(product_link, max_retries=3, backoff_factor=1.5):
    """
    Versão robusta da obtenção de detalhes de um produto.
    
    Args:
        product_link: URL do produto
        max_retries: Número máximo de tentativas em caso de falha
        backoff_factor: Fator de espera entre tentativas
        
    Returns:
        Dicionário com detalhes do vendedor
    """
    if not product_link:
        return {}
    
    # Rotacionar diferentes user agents para evitar bloqueios
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
    ]
    
    # Usar diferentes parsers como fallback
    parsers = [
        parse_product_details_standard,
        parse_product_details_alternative,
        parse_product_details_minimal
    ]
    
    result = {}
    attempt = 0
    
    while attempt < max_retries and not result:
        try:
            # Escolher um user agent aleatoriamente
            headers = {
                'User-Agent': random.choice(user_agents),
                'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
                'Cache-Control': 'no-cache',
                'Pragma': 'no-cache'
            }
            
            wait_time = backoff_factor ** attempt
            if attempt > 0:
                time.sleep(wait_time)
                
            response = requests.get(product_link, headers=headers, timeout=15)
            
            if response.status_code != 200:
                debug_print(f"Resposta HTTP {response.status_code} ao acessar detalhes do produto")
                attempt += 1
                continue
            
            # Tentar cada um dos parsers até encontrar resultados
            for parser_index, parser_func in enumerate(parsers):
                try:
                    parser_result = parser_func(response.text)
                    if parser_result and any(parser_result.values()):
                        debug_print(f"Detalhes do produto obtidos com sucesso (parser {parser_index+1})")
                        return parser_result
                except Exception as parser_error:
                    debug_print(f"Erro no parser de detalhes {parser_index+1}: {str(parser_error)}")
            
            # Se chegou aqui, nenhum parser encontrou resultados
            debug_print(f"Nenhum detalhe encontrado para o produto após tentar {len(parsers)} parsers")
            attempt += 1
            
        except requests.exceptions.Timeout:
            debug_print(f"Timeout ao obter detalhes do produto. Tentativa {attempt+1}/{max_retries}")
            attempt += 1
        except requests.exceptions.ConnectionError:
            debug_print(f"Erro de conexão ao obter detalhes do produto. Tentativa {attempt+1}/{max_retries}")
            attempt += 1
        except Exception as e:
            debug_print(f"Erro ao obter detalhes do produto: {str(e)}")
            attempt += 1
    
    # Dados padrão se todas as tentativas falharem
    return {
        'seller_level': "Não informado",
        'sales': 0,
        'rating': 0
    }

def parse_product_details_standard(html):
    """
    Parser padrão para detalhes do produto no Mercado Livre.
    
    Args:
        html: HTML da página do produto
        
    Returns:
        Dicionário com detalhes do vendedor
    """
    soup = BeautifulSoup(html, 'html.parser')
    
    # Nível do vendedor
    seller_level = "Não informado"
    seller_level_element = soup.find('span', class_='ui-seller-info__status-info')
    if seller_level_element:
        seller_level = seller_level_element.text
    
    # Vendas
    sales = 0
    sales_element = soup.find('strong', class_='ui-seller-info__sales-number')
    if sales_element:
        match = re.search(r'\d+', sales_element.text)
        if match:
            try:
                sales = int(match.group())
            except:
                pass
    
    # Avaliação
    rating = 0
    rating_element = soup.find('span', class_='ui-seller-info__rating-average')
    if rating_element:
        try:
            rating = float(rating_element.text.replace(',', '.'))
        except:
            pass
    
    return {
        'seller_level': seller_level,
        'sales': sales,
        'rating': rating
    }

def parse_product_details_alternative(html):
    """
    Parser alternativo para detalhes do produto no Mercado Livre.
    
    Args:
        html: HTML da página do produto
        
    Returns:
        Dicionário com detalhes do vendedor
    """
    soup = BeautifulSoup(html, 'html.parser')
    
    # Nível do vendedor (seletores alternativos)
    seller_level = "Não informado"
    seller_level_selectors = [
        '.seller-info__status-info',
        '.seller-info__status',
        '.ui-pdp-seller__label-title',
        '.ui-pdp-action-modal__link'
    ]
    
    for selector in seller_level_selectors:
        element = soup.select_one(selector)
        if element and element.text.strip():
            seller_level = element.text.strip()
            if any(status in seller_level.lower() for status in ['líder', 'platinum', 'gold', 'excelente', 'bom']):
                break
    
    # Vendas
    sales = 0
    sales_selectors = [
        '.ui-seller-info__sales-number',
        '.seller-info__sales-number',
        '.ui-pdp-seller__sales-description'
    ]
    
    for selector in sales_selectors:
        element = soup.select_one(selector)
        if element:
            match = re.search(r'\d+', element.text)
            if match:
                try:
                    sales = int(match.group())
                    break
                except:
                    pass
    
    # Avaliação
    rating = 0
    rating_selectors = [
        '.ui-seller-info__rating-average',
        '.seller-info__rating-average',
        '.ui-pdp-seller__reputation-score'
    ]
    
    for selector in rating_selectors:
        element = soup.select_one(selector)
        if element:
            try:
                rating_text = element.text.strip()
                rating = float(rating_text.replace(',', '.'))
                break
            except:
                pass
    
    # Se não encontrou rating específico, procurar por estrelas
    if rating == 0:
        stars_elements = soup.select('.ui-pdp-seller__reputation-stars .ui-pdp-icon--star-filled')
        if stars_elements:
            rating = len(stars_elements)
    
    return {
        'seller_level': seller_level,
        'sales': sales,
        'rating': rating
    }

def parse_product_details_minimal(html):
    """
    Parser minimalista para detalhes do produto.
    
    Args:
        html: HTML da página do produto
        
    Returns:
        Dicionário com detalhes do vendedor
    """
    # Usar expressões regulares para extrair informações diretamente do HTML
    
    # Nível do vendedor
    seller_level = "Não informado"
    seller_patterns = [
        r'MercadoL[íi]der\s*(Platinum|Gold)?',
        r'Vendedor\s*(Platinum|Gold)?',
        r'Reputação do vendedor[^<>]*?(\w+)',
        r'seller[^<>]*?level[^<>]*?(\w+)'
    ]
    
    for pattern in seller_patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            level = match.group(1) if match.group(1) else "Regular"
            seller_level = f"MercadoLíder {level}"
            break
    
    # Vendas
    sales = 0
    sales_patterns = [
        r'(\d+)\s*vendas',
        r'(\d+)\s*vendido',
        r'vendeu\s*(\d+)',
        r'sales[^<>]*?(\d+)'
    ]
    
    for pattern in sales_patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            try:
                sales = int(match.group(1))
                break
            except:
                pass
    
    # Avaliação
    rating = 0
    rating_patterns = [
        r'(\d[.,]\d+)\s*estrelas',
        r'rating[^<>]*?(\d[.,]\d+)',
        r'(\d)[.,](\d+)\s*/\s*5'
    ]
    
    for pattern in rating_patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            try:
                if match.group(2):  # Se capturou dois grupos (ex: 4,5)
                    rating = float(f"{match.group(1)}.{match.group(2)}")
                else:
                    rating = float(match.group(1).replace(',', '.'))
                break
            except:
                pass
    
    return {
        'seller_level': seller_level,
        'sales': sales,
        'rating': rating
    }

def calculate_mercado_livre_fees(price):
    """
    Calcula as taxas do Mercado Livre para um preço.
    
    Args:
        price: Preço do produto
        
    Returns:
        Dicionário com valores calculados
    """
    if not isinstance(price, (int, float)) or price <= 0:
        return {'price': 0, 'fee': 0, 'net': 0, 'margin': 0}
    
    # Taxa do Mercado Livre (aproximadamente 16%)
    fee_percentage = 0.16
    fee = price * fee_percentage
    net = price - fee
    margin = (net / price) * 100
    
    return {
        'price': price,
        'fee': fee,
        'net': net,
        'margin': margin
    }

def calculate_mercado_livre_fees_detailed(price, category=None):
    """
    Calcula as taxas do Mercado Livre para um preço com mais detalhes.
    
    Args:
        price: Preço do produto
        category: Categoria do produto (opcional, para taxas específicas)
        
    Returns:
        Dicionário com valores calculados
    """
    if not isinstance(price, (int, float)) or price <= 0:
        return {'price': 0, 'fee': 0, 'net': 0, 'margin': 0, 'details': {}}
    
    # Taxas base do Mercado Livre (variam por categoria)
    base_fee_percentage = 0.16  # 16% padrão
    
    # Ajuste de taxa por categoria (se fornecida)
    if category:
        category_lower = category.lower()
        
        # Categorias com taxas diferentes
        if 'celular' in category_lower or 'smartphone' in category_lower:
            base_fee_percentage = 0.17  # 17% para celulares
        elif 'informática' in category_lower or 'computador' in category_lower:
            base_fee_percentage = 0.16  # 16% para informática
        elif 'móveis' in category_lower or 'decoração' in category_lower:
            base_fee_percentage = 0.15  # 15% para móveis
    
    # Taxa de venda
    sale_fee = price * base_fee_percentage
    
    # Taxa fixa para produtos abaixo de R$79 (hipotético, verificar valores atuais)
    fixed_fee = 0
    if price < 79:
        fixed_fee = 5.0
    
    # Taxa de antifraude (estimativa)
    antifraud_fee = price * 0.015 if price > 120 else 0
    
    # Taxa de frete (estimativa)
    shipping_fee = 0  # Estimativa, varia por produto
    
    # Total de taxas
    total_fee = sale_fee + fixed_fee + antifraud_fee + shipping_fee
    
    # Valor líquido
    net = price - total_fee
    
    # Margem
    margin = (net / price) * 100 if price > 0 else 0
    
    return {
        'price': price,
        'fee': total_fee,
        'net': net,
        'margin': margin,
        'details': {
            'base_fee_percentage': base_fee_percentage * 100,
            'sale_fee': sale_fee,
            'fixed_fee': fixed_fee,
            'antifraud_fee': antifraud_fee,
            'shipping_fee': shipping_fee
        }
    }

# ------------------------------------------------------
# ANÁLISE COM IA
# ------------------------------------------------------

def analyze_product_with_ai(product_data, market_data, seller_data, fees_data, model="gpt-3.5-turbo"):
    """
    Analisa um produto usando IA.
    
    Args:
        product_data: Dados do produto
        market_data: Dados do mercado
        seller_data: Dados dos vendedores
        fees_data: Dados de taxas
        model: Modelo de IA a usar
        
    Returns:
        Dicionário com análise completa
    """
    # Verificar se temos API key e dados suficientes
    if not openai.api_key:
        return fallback_analysis(product_data, market_data, seller_data, fees_data)
    
    if not market_data:
        return fallback_analysis(product_data, market_data, seller_data, fees_data)
    
    try:
        # Preparar os dados para o prompt
        product_name = ""
        initial_price = None
        
        if isinstance(product_data, dict):
            product_name = product_data.get('description', '')
            initial_price = product_data.get('price')
        else:
            product_name = str(product_data)
        
        # Calcular médias
        prices = [p.get('price', 0) for p in market_data if p.get('price', 0) > 0]
        avg_price = sum(prices) / len(prices) if prices else 0
        
        sold_counts = [p.get('sold_count', 0) for p in market_data if p.get('sold_count', 0) > 0]
        avg_sold = sum(sold_counts) / len(sold_counts) if sold_counts else 0
        
        # Dados de taxas
        margin = 0
        if fees_data and 'margin' in fees_data:
            margin = fees_data['margin']
        
        # Dados de vendedores
        competition_level = 0
        if seller_data:
            high_level_count = 0
            for seller in seller_data:
                level = seller.get('seller_level', '')
                if any(status in level for status in ['Líder', 'Platinum', 'Gold']):
                    high_level_count += 1
            
            if len(seller_data) > 0:
                competition_level = (high_level_count / len(seller_data)) * 100
        
        # Criar o prompt para a IA
        prompt = f"""
        Analise este produto para venda no Mercado Livre:
        
        PRODUTO: {product_name}
        
        DADOS DE MERCADO:
        - Preço do produto no meu estoque: {initial_price if initial_price else 'Não disponível'}
        - Preço médio no Mercado Livre: R$ {avg_price:.2f}
        - Vendas médias: {avg_sold:.1f} unidades
        - Nível de competição: {competition_level:.1f}% dos vendedores são de alto nível
        - Margem após taxas do ML: {margin:.1f}%
        
        Forneça uma análise estruturada no seguinte formato JSON:
        {
            "price_analysis": {
                "score": [0-10],
                "average_price": {avg_price},
                "average_margin": {margin},
                "details": "Sua análise sobre preço e margens"
            },
            "competition_analysis": {
                "score": [0-10],
                "high_level_sellers": {competition_level},
                "details": "Sua análise sobre a concorrência"
            },
            "demand_analysis": {
                "score": [0-10],
                "average_sold": {avg_sold},
                "details": "Sua análise sobre a demanda"
            },
            "overall_score": [0-10],
            "recommendation": "Sua recomendação final"
        }
        
        Regras para pontuação:
        - Preço: Margens maiores que 85% são excelentes (10 pontos), abaixo de 70% são ruins (2 pontos)
        - Concorrência: Menos de 20% de vendedores de alto nível é excelente (10 pontos), mais de 80% é ruim (2 pontos)
        - Demanda: Mais de 1000 vendas é excelente (10 pontos), menos de 50 é ruim (2 pontos)
        - Overall score deve ser uma média ponderada (preço 30%, concorrência 30%, demanda 40%)
        - Recomendação deve ser "Altamente recomendado" (score >= 7), "Recomendado" (score >= 5), "Neutro" (score >= 3) ou "Não recomendado" (score < 3)
        
        Responda apenas com o JSON, sem texto adicional.
        """
        
        # Chamar a API da OpenAI
        response = openai.ChatCompletion.create(
            model=model,
            messages=[
                {"role": "system", "content": "Você é um especialista em análise de mercado para vendedores do Mercado Livre. Sua análise deve ser precisa, objetiva e orientada a resultados."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=800
        )
        
        # Extrair a resposta
        ai_response = response.choices[0].message['content']
        
        # Processar a resposta JSON
        return parse_ai_response(ai_response)
        
    except Exception as e:
        log_error("Erro ao usar IA para análise", e)
        return fallback_analysis(product_data, market_data, seller_data, fees_data)

def analyze_product_with_ai_enhanced(product_data, market_data, seller_data, fees_data, model="gpt-3.5-turbo", retries=2):
    """
    Versão melhorada da análise de produtos usando IA.
    
    Args:
        product_data: Dados do produto
        market_data: Dados do mercado
        seller_data: Dados dos vendedores
        fees_data: Dados de taxas
        model: Modelo de IA a usar
        retries: Número de tentativas em caso de falha
        
    Returns:
        Dicionário com análise completa
    """
    # Verificar se temos API key e dados suficientes
    if not openai.api_key:
        return fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data)
    
    if not market_data:
        return fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data)
    
    for attempt in range(retries + 1):
        try:
            # Extrair e processar dados do produto de forma mais robusta
            product_info = extract_product_info(product_data)
            market_metrics = calculate_market_metrics(market_data)
            seller_metrics = calculate_seller_metrics(seller_data)
            fee_metrics = process_fee_metrics(fees_data)
            
            # Criar um prompt mais estruturado e informativo
            prompt = create_enhanced_prompt(
                product_info,
                market_metrics,
                seller_metrics,
                fee_metrics
            )
            
            # Sistema de mensagens mais específico
            system_message = """Você é um especialista em análise de mercado para vendedores do Mercado Livre com anos de experiência.
            Sua análise deve ser precisa, objetiva e fornecer insights práticos que permitam ao vendedor tomar decisões informadas.
            Use os dados fornecidos para gerar uma análise aprofundada e prática, identificando oportunidades e riscos.
            Sua resposta deve ser estruturada exatamente no formato JSON solicitado, sem texto adicional."""
            
            # Chamar a API da OpenAI com gerenciamento de erros
            response = openai.ChatCompletion.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,  # Temperatura mais baixa para maior precisão
                max_tokens=1000,
                request_timeout=30  # Timeout maior
            )
            
            # Extrair a resposta
            ai_response = response.choices[0].message['content']
            
            # Processar e validar a resposta JSON
            analysis = parse_and_validate_ai_response(ai_response)
            
            if analysis:
                return enhance_analysis_with_trends(analysis, market_data, seller_data)
            else:
                # Se a análise falhou na validação e ainda temos tentativas
                if attempt < retries:
                    log_warning(f"Resposta da IA inválida, tentando novamente ({attempt+1}/{retries})")
                    time.sleep(2)  # Esperar um pouco antes de tentar novamente
                    continue
                else:
                    log_warning("Todas as tentativas de análise com IA falharam, usando método alternativo")
                    return fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data)
                
        except Exception as e:
            log_error(f"Erro ao usar IA para análise (tentativa {attempt+1}/{retries+1})", e)
            
            # Se ainda temos tentativas, tentar novamente
            if attempt < retries:
                time.sleep(2)  # Esperar um pouco antes de tentar novamente
                continue
            else:
                return fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data)
    
    # Não deveria chegar aqui, mas por segurança
    return fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data)

def extract_product_info(product_data):
    """
    Extrai informações do produto de forma robusta.
    
    Args:
        product_data: Dados do produto
        
    Returns:
        Dicionário com informações do produto
    """
    product_name = ""
    product_code = ""
    initial_price = None
    
    if isinstance(product_data, dict):
        product_name = product_data.get('description', '')
        product_code = product_data.get('code', '')
        initial_price = product_data.get('price')
    else:
        product_name = str(product_data)
    
    # Extrair características do produto
    product_type = classify_product_type(product_name)
    
    return {
        'name': product_name,
        'code': product_code,
        'initial_price': initial_price,
        'type': product_type
    }

def classify_product_type(product_name):
    """
    Tenta classificar o tipo de produto com base no nome.
    
    Args:
        product_name: Nome do produto
        
    Returns:
        Tipo do produto ou "Diversos"
    """
    product_name_lower = product_name.lower()
    
    # Dicionário de categorias e palavras-chave
    categories = {
        'Eletrônicos': ['celular', 'smartphone', 'tv', 'televisão', 'monitor', 'tablet', 'notebook', 'laptop', 'fone', 'headphone'],
        'Informática': ['computador', 'pc', 'teclado', 'mouse', 'impressora', 'scanner', 'webcam', 'hd', 'ssd', 'pendrive'],
        'Móveis': ['mesa', 'cadeira', 'sofá', 'poltrona', 'armário', 'estante', 'cama', 'guarda-roupa', 'criado-mudo'],
        'Eletrodomésticos': ['geladeira', 'fogão', 'microondas', 'liquidificador', 'batedeira', 'cafeteira', 'aspirador'],
        'Ferramentas': ['martelo', 'chave', 'parafusadeira', 'furadeira', 'alicate', 'serra', 'esmerilhadeira'],
        'Decoração': ['tapete', 'cortina', 'quadro', 'luminária', 'espelho', 'vaso', 'almofada'],
        'Vestuário': ['camisa', 'camiseta', 'calça', 'vestido', 'bermuda', 'jaqueta', 'casaco', 'sapato', 'tênis'],
        'Brinquedos': ['boneca', 'carrinho', 'jogo', 'puzzle', 'quebra-cabeça', 'lego', 'nerf']
    }
    
    for category, keywords in categories.items():
        if any(keyword in product_name_lower for keyword in keywords):
            return category
    
    return "Diversos"

def calculate_market_metrics(market_data):
    """
    Calcula métricas de mercado de forma mais abrangente.
    
    Args:
        market_data: Dados do mercado
        
    Returns:
        Dicionário com métricas de mercado
    """
    # Calcular dados básicos
    prices = [p.get('price', 0) for p in market_data if p.get('price', 0) > 0]
    
    # Metricas de preço
    price_metrics = {
        'avg_price': sum(prices) / len(prices) if prices else 0,
        'min_price': min(prices) if prices else 0,
        'max_price': max(prices) if prices else 0,
        'price_range': max(prices) - min(prices) if prices else 0,
        'price_std_dev': calculate_std_dev(prices)
    }
    
    # Vendas
    sold_counts = [p.get('sold_count', 0) for p in market_data if p.get('sold_count', 0) > 0]
    
    sales_metrics = {
        'avg_sold': sum(sold_counts) / len(sold_counts) if sold_counts else 0,
        'min_sold': min(sold_counts) if sold_counts else 0,
        'max_sold': max(sold_counts) if sold_counts else 0,
        'total_competitors': len(market_data)
    }
    
    # Classificar a variação de preço
    price_variation = "Baixa"
    if price_metrics['price_std_dev'] > price_metrics['avg_price'] * 0.2:
        price_variation = "Alta"
    elif price_metrics['price_std_dev'] > price_metrics['avg_price'] * 0.1:
        price_variation = "Média"
    
    # Classificar a demanda
    demand_level = "Baixa"
    if sales_metrics['avg_sold'] >= 500:
        demand_level = "Alta"
    elif sales_metrics['avg_sold'] >= 100:
        demand_level = "Média"
    
    return {
        'price': price_metrics,
        'sales': sales_metrics,
        'price_variation': price_variation,
        'demand_level': demand_level
    }

def calculate_std_dev(values):
    """
    Calcula o desvio padrão de uma lista de valores.
    
    Args:
        values: Lista de valores
        
    Returns:
        Desvio padrão
    """
    if not values or len(values) < 2:
        return 0
    
    mean = sum(values) / len(values)
    variance = sum((x - mean) ** 2 for x in values) / len(values)
    return variance ** 0.5

def calculate_seller_metrics(seller_data):
    """
    Calcula métricas dos vendedores de forma mais detalhada.
    
    Args:
        seller_data: Dados dos vendedores
        
    Returns:
        Dicionário com métricas dos vendedores
    """
    if not seller_data:
        return {
            'high_level_percent': 0,
            'avg_rating': 0,
            'avg_sales': 0,
            'competition_level': "Desconhecido"
        }
    
    # Calcular percentual de vendedores de alto nível
    high_level_count = 0
    for seller in seller_data:
        level = seller.get('seller_level', '')
        if any(status in level.lower() for status in ['líder', 'platinum', 'gold', 'excelente']):
            high_level_count += 1
    
    high_level_percent = (high_level_count / len(seller_data)) * 100 if seller_data else 0
    
    # Calcular avaliação média
    ratings = [seller.get('rating', 0) for seller in seller_data if seller.get('rating', 0) > 0]
    avg_rating = sum(ratings) / len(ratings) if ratings else 0
    
    # Calcular vendas médias
    sales = [seller.get('sales', 0) for seller in seller_data if seller.get('sales', 0) > 0]
    avg_sales = sum(sales) / len(sales) if sales else 0
    
    # Classificar o nível de competição
    competition_level = "Baixo"
    if high_level_percent >= 70:
        competition_level = "Alto"
    elif high_level_percent >= 40:
        competition_level = "Médio"
    
    # Ajustar com base na avaliação média
    if avg_rating >= 4.7:
        # Se a avaliação média for muito alta, a competição é mais forte
        if competition_level == "Médio":
            competition_level = "Alto"
    
    return {
        'high_level_percent': high_level_percent,
        'avg_rating': avg_rating,
        'avg_sales': avg_sales,
        'competition_level': competition_level
    }

def process_fee_metrics(fees_data):
    """
    Processa métricas de taxas de forma mais detalhada.
    
    Args:
        fees_data: Dados de taxas
        
    Returns:
        Dicionário com métricas de taxas
    """
    if not fees_data:
        return {
            'margin': 84,  # Valor padrão
            'fee_percentage': 16,
            'net_revenue_ratio': 0.84,
            'profitability': "Desconhecido"
        }
    
    # Extrair margem e taxas
    margin = fees_data.get('margin', 84)
    fee = fees_data.get('fee', 0)
    price = fees_data.get('price', 0)
    
    # Calcular percentual de taxa
    fee_percentage = (fee / price) * 100 if price > 0 else 16
    
    # Calcular proporção de receita líquida
    net_revenue_ratio = 1 - (fee_percentage / 100)
    
    # Classificar a rentabilidade
    profitability = "Média"
    if margin >= 85:
        profitability = "Alta"
    elif margin >= 80:
        profitability = "Boa"
    elif margin >= 75:
        profitability = "Média"
    elif margin >= 70:
        profitability = "Baixa"
    else:
        profitability = "Muito baixa"
    
    return {
        'margin': margin,
        'fee_percentage': fee_percentage,
        'net_revenue_ratio': net_revenue_ratio,
        'profitability': profitability
    }

def create_enhanced_prompt(product_info, market_metrics, seller_metrics, fee_metrics):
    """
    Cria um prompt melhorado para a IA.
    
    Args:
        product_info: Informações do produto
        market_metrics: Métricas de mercado
        seller_metrics: Métricas dos vendedores
        fee_metrics: Métricas de taxas
        
    Returns:
        Prompt formatado
    """
    return f"""
    Analise este produto para venda no Mercado Livre como um especialista em e-commerce:
    
    ===== PRODUTO =====
    Nome: {product_info['name']}
    Tipo: {product_info['type']}
    Preço no estoque: {product_info['initial_price'] if product_info['initial_price'] else 'Não disponível'}
    
    ===== DADOS DE MERCADO =====
    PREÇOS:
    - Preço médio no ML: R$ {market_metrics['price']['avg_price']:.2f}
    - Preço mínimo: R$ {market_metrics['price']['min_price']:.2f}
    - Preço máximo: R$ {market_metrics['price']['max_price']:.2f}
    - Variação de preço: {market_metrics['price_variation']} (Desvio: {market_metrics['price']['price_std_dev']:.2f})
    
    DEMANDA:
    - Vendas médias: {market_metrics['sales']['avg_sold']:.1f} unidades
    - Vendas máximas: {market_metrics['sales']['max_sold']}
    - Nível de demanda: {market_metrics['demand_level']}
    - Total de concorrentes: {market_metrics['sales']['total_competitors']}
    
    CONCORRÊNCIA:
    - Porcentagem de vendedores de alto nível: {seller_metrics['high_level_percent']:.1f}%
    - Avaliação média dos vendedores: {seller_metrics['avg_rating']:.1f}/5
    - Vendas médias por vendedor: {seller_metrics['avg_sales']:.0f}
    - Nível de competição: {seller_metrics['competition_level']}
    
    FINANCEIRO:
    - Margem após taxas: {fee_metrics['margin']:.1f}%
    - Taxa do ML: {fee_metrics['fee_percentage']:.1f}%
    - Proporção de receita líquida: {fee_metrics['net_revenue_ratio']:.2f}
    - Classificação de rentabilidade: {fee_metrics['profitability']}
    
    ===== FORMATO DA RESPOSTA =====
    Forneça uma análise estruturada no seguinte formato JSON:
    {{
        "price_analysis": {{
            "score": [0-10],
            "average_price": {market_metrics['price']['avg_price']},
            "average_margin": {fee_metrics['margin']},
            "details": "Sua análise sobre preço e margens"
        }},
        "competition_analysis": {{
            "score": [0-10],
            "high_level_sellers": {seller_metrics['high_level_percent']},
            "details": "Sua análise sobre a concorrência"
        }},
        "demand_analysis": {{
            "score": [0-10],
            "average_sold": {market_metrics['sales']['avg_sold']},
            "details": "Sua análise sobre a demanda"
        }},
        "overall_score": [0-10],
        "recommendation": "Sua recomendação final",
        "improvement_suggestions": ["Sugestão 1", "Sugestão 2", "Sugestão 3"]
    }}
    
    Regras para pontuação:
    - Preço: Margens maiores que 85% são excelentes (10 pontos), abaixo de 70% são ruins (2 pontos)
    - Concorrência: Menos de 20% de vendedores de alto nível é excelente (10 pontos), mais de 80% é ruim (2 pontos)
    - Demanda: Mais de 1000 vendas é excelente (10 pontos), menos de 50 é ruim (2 pontos)
    - Overall score deve ser uma média ponderada (preço 30%, concorrência 30%, demanda 40%)
    - Recomendação deve ser uma das seguintes: "Altamente recomendado" (score >= 7), "Recomendado" (score >= 5), "Neutro" (score >= 3) ou "Não recomendado" (score < 3)
    - Inclua também 2-3 sugestões práticas para melhorar a competitividade deste produto no campo "improvement_suggestions"
    
    Responda apenas com o JSON, sem texto adicional.
    """

def parse_ai_response(response_text):
    """
    Processa a resposta da IA para extrair o JSON.
    
    Args:
        response_text: Texto da resposta da IA
        
    Returns:
        Dicionário com a análise
    """
    if not response_text:
        return get_default_analysis()
    
    try:
        # Limpar o texto para obter apenas o JSON
        cleaned_text = response_text.strip()
        
        # Remover marcadores de código se presentes
        if cleaned_text.startswith('```json'):
            cleaned_text = cleaned_text[7:]
        elif cleaned_text.startswith('```'):
            cleaned_text = cleaned_text[3:]
            
        if cleaned_text.endswith('```'):
            cleaned_text = cleaned_text[:-3]
            
        # Analisar o JSON
        analysis = json.loads(cleaned_text)
        
        # Garantir que todos os campos necessários estejam presentes
        required_fields = [
            'price_analysis', 'competition_analysis', 'demand_analysis', 
            'overall_score', 'recommendation'
        ]
        
        for field in required_fields:
            if field not in analysis:
                if field.endswith('_analysis'):
                    analysis[field] = {
                        'score': 5,
                        'details': f"Campo {field} não disponível"
                    }
                    
                    # Adicionar campos específicos
                    if field == 'price_analysis':
                        analysis[field]['average_price'] = 0
                        analysis[field]['average_margin'] = 0
                    elif field == 'competition_analysis':
                        analysis[field]['high_level_sellers'] = 0
                    elif field == 'demand_analysis':
                        analysis[field]['average_sold'] = 0
                elif field == 'overall_score':
                    analysis[field] = 5
                elif field == 'recommendation':
                    analysis[field] = "Neutro"
        
        return analysis
        
    except Exception as e:
        debug_print(f"Erro ao processar resposta da IA: {str(e)}")
        return get_default_analysis()

def parse_and_validate_ai_response(response_text):
    """
    Processa e valida a resposta da IA.
    
    Args:
        response_text: Texto da resposta da IA
        
    Returns:
        Dicionário com a análise ou None se inválido
    """
    if not response_text:
        return None
    
    try:
        # Limpar o texto para obter apenas o JSON
        cleaned_text = response_text.strip()
        
        # Remover marcadores de código se presentes
        if cleaned_text.startswith('```json'):
            cleaned_text = cleaned_text[7:]
        elif cleaned_text.startswith('```'):
            cleaned_text = cleaned_text[3:]
            
        if cleaned_text.endswith('```'):
            cleaned_text = cleaned_text[:-3]
            
        # Analisar o JSON
        analysis = json.loads(cleaned_text)
        
        # Validar campos obrigatórios
        required_fields = [
            'price_analysis', 'competition_analysis', 'demand_analysis', 
            'overall_score', 'recommendation'
        ]
        
        for field in required_fields:
            if field not in analysis:
                debug_print(f"Campo obrigatório ausente na resposta da IA: {field}")
                return None
            
            # Validar campos aninhados
            if field.endswith('_analysis'):
                if not isinstance(analysis[field], dict):
                    debug_print(f"Campo {field} não é um objeto")
                    return None
                
                if 'score' not in analysis[field]:
                    debug_print(f"Campo 'score' ausente em {field}")
                    return None
                
                if 'details' not in analysis[field]:
                    debug_print(f"Campo 'details' ausente em {field}")
                    return None
        
        # Validar score geral
        if not isinstance(analysis['overall_score'], (int, float)):
            debug_print("Score geral não é um número")
            return None
        
        if not 0 <= analysis['overall_score'] <= 10:
            debug_print(f"Score geral fora do intervalo: {analysis['overall_score']}")
            return None
        
        # Validar recomendação
        valid_recommendations = ["Altamente recomendado", "Recomendado", "Neutro", "Não recomendado"]
        if analysis['recommendation'] not in valid_recommendations:
            debug_print(f"Recomendação inválida: {analysis['recommendation']}")
            # Corrigir em vez de rejeitar
            score = analysis['overall_score']
            if score >= 7:
                analysis['recommendation'] = "Altamente recomendado"
            elif score >= 5:
                analysis['recommendation'] = "Recomendado"
            elif score >= 3:
                analysis['recommendation'] = "Neutro"
            else:
                analysis['recommendation'] = "Não recomendado"
        
        # Adicionar campo de sugestões se não existir
        if 'improvement_suggestions' not in analysis:
            analysis['improvement_suggestions'] = [
                "Pesquise os concorrentes para otimizar seu preço",
                "Melhore as imagens e descrição do produto",
                "Considere oferecer frete grátis para aumentar a conversão"
            ]
        
        return analysis
        
    except json.JSONDecodeError as e:
        debug_print(f"Erro ao decodificar JSON da resposta da IA: {str(e)}")
        return None
    except Exception as e:
        debug_print(f"Erro ao processar resposta da IA: {str(e)}")
        return None

def enhance_analysis_with_trends(analysis, market_data, seller_data):
    """
    Melhora a análise adicionando tendências e insights adicionais.
    
    Args:
        analysis: Análise base
        market_data: Dados do mercado
        seller_data: Dados dos vendedores
        
    Returns:
        Análise melhorada
    """
    # Adicionar campo de tendências se não existir
    if 'trends' not in analysis:
        analysis['trends'] = {}
    
    # Analisar tendência de preço
    prices = [p.get('price', 0) for p in market_data if p.get('price', 0) > 0]
    if len(prices) >= 5:
        price_trend = "Estável"
        price_variance = calculate_std_dev(prices) / (sum(prices) / len(prices)) if sum(prices) > 0 else 0
        
        if price_variance > 0.2:
            price_trend = "Volátil"
        
        analysis['trends']['price_trend'] = price_trend
        analysis['trends']['price_variance'] = price_variance
    
    # Analisar sazonalidade (simulado, em produção usaria dados históricos)
    current_month = datetime.now().month
    season_map = {
        # Verão
        1: "Alta",
        2: "Alta",
        
        # Outono
        3: "Média",
        4: "Média",
        5: "Baixa",
        
        # Inverno
        6: "Baixa",
        7: "Baixa",
        
        # Primavera
        8: "Baixa",
        9: "Média",
        
        # Pré-temporada de festas
        10: "Média",
        
        # Temporada de festas
        11: "Alta",
        12: "Alta"
    }
    
    analysis['trends']['seasonal_demand'] = season_map.get(current_month, "Média")
    
    # Adicionar métricas de qualidade dos concorrentes
    if seller_data:
        ratings = [seller.get('rating', 0) for seller in seller_data if seller.get('rating', 0) > 0]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0
        
        quality_assessment = "Média"
        if avg_rating >= 4.8:
            quality_assessment = "Excelente"
        elif avg_rating >= 4.5:
            quality_assessment = "Boa"
        elif avg_rating >= 4.0:
            quality_assessment = "Média"
        else:
            quality_assessment = "Baixa"
        
        analysis['trends']['competitor_quality'] = quality_assessment
    
    return analysis

def get_default_analysis():
    """
    Retorna uma análise padrão quando há falhas.
    
    Returns:
        Análise padrão
    """
    return {
        "price_analysis": {
            "score": 5,
            "average_price": 0,
            "average_margin": 0,
            "details": "Informações de preço não disponíveis"
        },
        "competition_analysis": {
            "score": 5,
            "high_level_sellers": 0,
            "details": "Informações de concorrência não disponíveis"
        },
        "demand_analysis": {
            "score": 5,
            "average_sold": 0,
            "details": "Informações de demanda não disponíveis"
        },
        "overall_score": 5,
        "recommendation": "Neutro (análise indisponível)"
    }

def fallback_analysis(product_data, market_data, seller_data, fees_data):
    """
    Análise alternativa quando a IA falha.
    
    Args:
        product_data: Dados do produto
        market_data: Dados do mercado
        seller_data: Dados dos vendedores
        fees_data: Dados de taxas
        
    Returns:
        Dicionário com análise
    """
    # Cálculos de preço
    prices = [p.get('price', 0) for p in market_data if p.get('price', 0) > 0]
    avg_price = sum(prices) / len(prices) if prices else 0
    
    # Margem
    margin = 0
    if fees_data and 'margin' in fees_data:
        margin = fees_data['margin']
    else:
        margin = 84  # Estimativa padrão
    
    # Pontuação de preço
    if margin >= 85:
        price_score = 10
        price_details = "Excelente margem"
    elif margin >= 80:
        price_score = 8
        price_details = "Boa margem"
    elif margin >= 75:
        price_score = 6
        price_details = "Margem razoável"
    elif margin >= 70:
        price_score = 4
        price_details = "Margem baixa"
    else:
        price_score = 2
        price_details = "Margem muito baixa"
    
    # Concorrência
    competition_level = 0
    if seller_data:
        high_level_count = 0
        for seller in seller_data:
            level = seller.get('seller_level', '')
            if any(status in level for status in ['Líder', 'Platinum', 'Gold']):
                high_level_count += 1
        
        if len(seller_data) > 0:
            competition_level = (high_level_count / len(seller_data)) * 100
    
    # Pontuação de concorrência
    if competition_level <= 20:
        competition_score = 10
        competition_details = "Concorrência muito baixa"
    elif competition_level <= 40:
        competition_score = 8
        competition_details = "Concorrência baixa"
    elif competition_level <= 60:
        competition_score = 6
        competition_details = "Concorrência moderada"
    elif competition_level <= 80:
        competition_score = 4
        competition_details = "Concorrência alta"
    else:
        competition_score = 2
        competition_details = "Concorrência muito alta"
    
    # Demanda
    sold_counts = [p.get('sold_count', 0) for p in market_data if p.get('sold_count', 0) > 0]
    avg_sold = sum(sold_counts) / len(sold_counts) if sold_counts else 0
    
    # Pontuação de demanda
    if avg_sold >= 1000:
        demand_score = 10
        demand_details = "Demanda extremamente alta"
    elif avg_sold >= 500:
        demand_score = 8
        demand_details = "Demanda alta"
    elif avg_sold >= 200:
        demand_score = 6
        demand_details = "Demanda moderada"
    elif avg_sold >= 50:
        demand_score = 4
        demand_details = "Demanda baixa"
    else:
        demand_score = 2
        demand_details = "Demanda muito baixa"
    
    # Pontuação geral
    overall_score = (price_score * 0.3) + (competition_score * 0.3) + (demand_score * 0.4)
    
    # Recomendação
    if overall_score >= 7:
        recommendation = "Altamente recomendado"
    elif overall_score >= 5:
        recommendation = "Recomendado"
    elif overall_score >= 3:
        recommendation = "Neutro"
    else:
        recommendation = "Não recomendado"
    
    return {
        "price_analysis": {
            "score": price_score,
            "average_price": avg_price,
            "average_margin": margin,
            "details": price_details
        },
        "competition_analysis": {
            "score": competition_score,
            "high_level_sellers": competition_level,
            "details": competition_details
        },
        "demand_analysis": {
            "score": demand_score,
            "average_sold": avg_sold,
            "details": demand_details
        },
        "overall_score": overall_score,
        "recommendation": recommendation
    }

def fallback_analysis_enhanced(product_data, market_data, seller_data, fees_data):
    """
    Versão melhorada da análise alternativa quando a IA falha.
    
    Args:
        product_data: Dados do produto
        market_data: Dados do mercado
        seller_data: Dados dos vendedores
        fees_data: Dados de taxas
        
    Returns:
        Dicionário com análise melhorada
    """
    # Extrair e processar dados
    product_info = extract_product_info(product_data)
    market_metrics = calculate_market_metrics(market_data)
    seller_metrics = calculate_seller_metrics(seller_data)
    fee_metrics = process_fee_metrics(fees_data)
    
    # Calcular pontuações
    # Pontuação de preço (baseada na margem)
    margin = fee_metrics['margin']
    
    if margin >= 85:
        price_score = 10
        price_details = "Excelente margem, muito acima da média do mercado."
    elif margin >= 80:
        price_score = 8
        price_details = "Boa margem, acima da média do mercado."
    elif margin >= 75:
        price_score = 6
        price_details = "Margem razoável, na média do mercado."
    elif margin >= 70:
        price_score = 4
        price_details = "Margem baixa, abaixo da média do mercado."
    else:
        price_score = 2
        price_details = "Margem muito baixa, significativamente abaixo do mercado."
    
    # Pontuação de concorrência
    competition_level = seller_metrics['high_level_percent']
    
    if competition_level <= 20:
        competition_score = 10
        competition_details = "Concorrência muito baixa, com poucos vendedores estabelecidos."
    elif competition_level <= 40:
        competition_score = 8
        competition_details = "Concorrência baixa, bom cenário para novos vendedores."
    elif competition_level <= 60:
        competition_score = 6
        competition_details = "Concorrência moderada, típica de produtos estabelecidos."
    elif competition_level <= 80:
        competition_score = 4
        competition_details = "Concorrência alta, dominada por vendedores experientes."
    else:
        competition_score = 2
        competition_details = "Concorrência muito alta, mercado saturado com vendedores de elite."
    
    # Pontuação de demanda
    avg_sold = market_metrics['sales']['avg_sold']
    
    if avg_sold >= 1000:
        demand_score = 10
        demand_details = "Demanda extremamente alta, produto muito procurado no mercado."
    elif avg_sold >= 500:
        demand_score = 8
        demand_details = "Demanda alta, produto com bom volume de vendas."
    elif avg_sold >= 200:
        demand_score = 6
        demand_details = "Demanda moderada, volume de vendas médio."
    elif avg_sold >= 50:
        demand_score = 4
        demand_details = "Demanda baixa, poucas vendas registradas."
    else:
        demand_score = 2
        demand_details = "Demanda muito baixa, produto com pouquíssimas vendas."
    
    # Pontuação geral (média ponderada)
    overall_score = (price_score * 0.3) + (competition_score * 0.3) + (demand_score * 0.4)
    
    # Recomendação
    if overall_score >= 7:
        recommendation = "Altamente recomendado"
    elif overall_score >= 5:
        recommendation = "Recomendado"
    elif overall_score >= 3:
        recommendation = "Neutro"
    else:
        recommendation = "Não recomendado"
    
    # Sugestões de melhoria baseadas nos pontos fracos
    improvement_suggestions = []
    
    # Adicionar sugestões específicas baseadas nas pontuações mais baixas
    lowest_score = min(price_score, competition_score, demand_score)
    
    if lowest_score == price_score:
        if price_score <= 4:
            improvement_suggestions.append("Otimize seu preço para melhorar a margem, considerando a média do mercado de R$ " + f"{market_metrics['price']['avg_price']:.2f}")
            improvement_suggestions.append("Busque fornecedores alternativos para reduzir o custo de aquisição")
    
    if lowest_score == competition_score:
        if competition_score <= 4:
            improvement_suggestions.append("Destaque-se dos concorrentes com descrições mais detalhadas e fotos de melhor qualidade")
            improvement_suggestions.append("Ofereça benefícios exclusivos como frete grátis ou brindes")
    
    if lowest_score == demand_score:
        if demand_score <= 4:
            improvement_suggestions.append("Invista em anúncios patrocinados para aumentar a visibilidade do produto")
            improvement_suggestions.append("Considere pacotes promocionais ou kits para aumentar o ticket médio")
    
    # Garantir pelo menos 3 sugestões
    if len(improvement_suggestions) < 3:
        general_suggestions = [
            "Mantenha um bom tempo de resposta para aumentar a reputação no Mercado Livre",
            "Acompanhe regularmente os preços dos concorrentes e ajuste sua estratégia",
            "Invista em uma descrição completa do produto, com especificações detalhadas",
            "Utilize imagens de alta qualidade e em diferentes ângulos",
            "Ofereça garantia estendida para aumentar a confiança do comprador"
        ]
        
        for suggestion in general_suggestions:
            if suggestion not in improvement_suggestions:
                improvement_suggestions.append(suggestion)
                if len(improvement_suggestions) >= 3:
                    break
    
    # Construir análise completa
    analysis = {
        "price_analysis": {
            "score": price_score,
            "average_price": market_metrics['price']['avg_price'],
            "average_margin": margin,
            "details": price_details
        },
        "competition_analysis": {
            "score": competition_score,
            "high_level_sellers": competition_level,
            "details": competition_details
        },
        "demand_analysis": {
            "score": demand_score,
            "average_sold": avg_sold,
            "details": demand_details
        },
        "overall_score": overall_score,
        "recommendation": recommendation,
        "improvement_suggestions": improvement_suggestions[:3],  # Limitar a 3 sugestões
        "trends": {
            "price_trend": market_metrics['price_variation'],
            "competitor_quality": seller_metrics['competition_level'],
            "seasonal_demand": "Média"  # Valor padrão
        }
    }
    
    return analysis

# ------------------------------------------------------
# GERAÇÃO DE KITS E RECOMENDAÇÕES
# ------------------------------------------------------

def generate_kit_recommendations(product_analyses, max_kits=5, kit_size=3, model="gpt-3.5-turbo"):
    """
    Gera recomendações de kits de produtos usando IA.
    
    Args:
        product_analyses: Lista de análises de produtos
        max_kits: Número máximo de kits a gerar
        kit_size: Tamanho de cada kit
        model: Modelo de IA a usar
        
    Returns:
        Lista de kits recomendados
    """
    # Filtrar produtos válidos
    valid_products = [p for p in product_analyses if p.get('found', True)]
    
    # Verificar se temos produtos suficientes
    if len(valid_products) < kit_size:
        log_warning(f"Produtos insuficientes para gerar kits ({len(valid_products)})")
        return []
    
    # Se não temos API key, usar método alternativo
    if not openai.api_key:
        return generate_traditional_kits(product_analyses, max_kits, kit_size)
    
    try:
        # Ordenar produtos por pontuação
        sorted_products = sorted(
            valid_products, 
            key=lambda x: x.get('overall_score', 0), 
            reverse=True
        )
        
        # Obter os top produtos
        top_products = sorted_products[:15]  # Limitar a 15 para o prompt
        
        # Preparar informações dos produtos para o prompt
        product_infos = []
        for i, product in enumerate(top_products):
            name = product.get('product_name', f"Produto {i+1}")
            price = product.get('price_analysis', {}).get('average_price', 0)
            score = product.get('overall_score', 0)
            sold = product.get('demand_analysis', {}).get('average_sold', 0)
            
            product_infos.append(f"ID: {i+1}, Nome: {name}, Preço: R$ {price:.2f}, Score: {score:.1f}, Vendas: {sold:.0f}")
        
        products_text = "\n".join(product_infos)
        
        # Criar prompt para a IA
        prompt = f"""
        Com base na lista de produtos abaixo, recomende {max_kits} kits diferentes para venda no Mercado Livre.
        
        PRODUTOS DISPONÍVEIS:
        {products_text}
        
        Para cada kit:
        1. Selecione {kit_size} produtos que combinem bem
        2. Calcule o preço individual total e aplique um desconto apropriado (5-10%)
        3. Explique brevemente por que este kit seria atrativo para os clientes
        
        Forneça suas recomendações em formato JSON conforme abaixo:
        
        [
            {{
                "kit_name": "Nome do Kit 1",
                "products": [IDs dos produtos, ex: [1, 5, 7]],
                "individual_prices": [preços individuais],
                "total_price": preço total sem desconto,
                "kit_price": preço com desconto,
                "discount": percentual de desconto,
                "average_score": pontuação média do kit,
                "reasoning": "Breve explicação de por que esse kit funcionaria bem"
            }},
            // mais kits...
        ]
        
        Diretrizes para criação de kits:
        - Considere produtos complementares que fazem sentido juntos
        - Priorize produtos com maior pontuação e demanda
        - Crie variedade entre os kits (não repita os mesmos produtos em todos)
        - Pense em diferentes tipos de clientes e necessidades
        
        Responda apenas com o JSON, sem texto adicional.
        """
        
        # Chamar a API da OpenAI
        response = openai.ChatCompletion.create(
            model=model,
            messages=[
                {"role": "system", "content": "Você é um especialista em estratégia de vendas para o Mercado Livre, especializado em criar kits de produtos que maximizam as vendas e rentabilidade."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=1200
        )
        
        # Extrair a resposta
        ai_response = response.choices[0].message['content']
        
        # Processar a resposta JSON
        kits = parse_kit_recommendations(ai_response, top_products)
        
        if kits:
            log_success(f"Gerados {len(kits)} kits com IA")
            return kits
        else:
            log_warning("Falha ao gerar kits com IA, usando método alternativo")
            return generate_traditional_kits(product_analyses, max_kits, kit_size)
            
    except Exception as e:
        log_error("Erro ao gerar kits com IA", e)
        return generate_traditional_kits(product_analyses, max_kits, kit_size)

def parse_kit_recommendations(response_text, products):
    """
    Processa a resposta da IA para extrair kits.
    
    Args:
        response_text: Texto da resposta da IA
        products: Lista de produtos disponíveis
        
    Returns:
        Lista de kits recomendados
    """
    if not response_text:
        return []
    
    try:
        # Limpar o texto para obter apenas o JSON
        cleaned_text = response_text.strip()
        
        # Remover marcadores de código se presentes
        if cleaned_text.startswith('```json'):
            cleaned_text = cleaned_text[7:]
        elif cleaned_text.startswith('```'):
            cleaned_text = cleaned_text[3:]
            
        if cleaned_text.endswith('```'):
            cleaned_text = cleaned_text[:-3]
            
        # Analisar o JSON
        kits_data = json.loads(cleaned_text)
        
        # Verificar se é uma lista
        if not isinstance(kits_data, list):
            return []
        
        # Processar cada kit
        processed_kits = []
        
        for kit_data in kits_data:
            # Extrair IDs dos produtos
            product_ids = kit_data.get('products', [])
            product_names = []
            
            # Converter IDs para nomes de produtos
            for pid in product_ids:
                if isinstance(pid, int) and 1 <= pid <= len(products):
                    product_names.append(products[pid-1].get('product_name', f"Produto {pid}"))
                else:
                    product_names.append(f"Produto {pid}")
            
            # Preços individuais
            individual_prices = kit_data.get('individual_prices', [])
            
            # Se não tiver preços ou o número for diferente, recalcular
            if not individual_prices or len(individual_prices) != len(product_ids):
                individual_prices = []
                for pid in product_ids:
                    if isinstance(pid, int) and 1 <= pid <= len(products):
                        price = products[pid-1].get('price_analysis', {}).get('average_price', 0)
                        individual_prices.append(price)
                    else:
                        individual_prices.append(0)
            
            # Calcular total e preço com desconto
            total_price = kit_data.get('total_price', sum(individual_prices))
            discount = kit_data.get('discount', 5)
            kit_price = kit_data.get('kit_price', total_price * (1 - discount/100))
            
            # Pontuação média
            scores = []
            for pid in product_ids:
                if isinstance(pid, int) and 1 <= pid <= len(products):
                    scores.append(products[pid-1].get('overall_score', 0))
            
            avg_score = kit_data.get('average_score', sum(scores) / len(scores) if scores else 0)
            
            # Criar kit processado
            processed_kit = {
                'kit_name': kit_data.get('kit_name', f"Kit {len(processed_kits) + 1}"),
                'products': product_names,
                'individual_prices': individual_prices,
                'total_price': total_price,
                'kit_price': kit_price,
                'discount': discount,
                'average_score': avg_score,
                'reasoning': kit_data.get('reasoning', "Kit com produtos complementares")
            }
            
            processed_kits.append(processed_kit)
        
        return processed_kits
        
    except Exception as e:
        debug_print(f"Erro ao processar kits da IA: {str(e)}")
        return []

def generate_traditional_kits(product_analyses, max_kits=5, kit_size=3):
    """
    Gera kits de forma tradicional (sem IA).
    
    Args:
        product_analyses: Lista de análises de produtos
        max_kits: Número máximo de kits a gerar
        kit_size: Tamanho de cada kit
        
    Returns:
        Lista de kits gerados
    """
    # Filtrar produtos válidos e recomendados
    good_products = [p for p in product_analyses 
                   if p.get('found', True) and p.get('overall_score', 0) >= 5]
    
    # Se não tiver produtos suficientes, baixar o critério
    if len(good_products) < kit_size:
        good_products = [p for p in product_analyses 
                       if p.get('found', True) and p.get('overall_score', 0) >= 3]
    
    # Verificar se temos produtos suficientes
    if len(good_products) < kit_size:
        return []
    
    # Ordenar por pontuação
    sorted_products = sorted(good_products, key=lambda x: x.get('overall_score', 0), reverse=True)
    
    # Gerar kits
    kits = []
    max_possible_kits = min(max_kits, len(sorted_products) - kit_size + 1)
    
    for i in range(max_possible_kits):
        # Selecionar produtos para o kit
        kit_products = sorted_products[i:i+kit_size]
        
        # Extrair nomes e preços
        product_names = []
        individual_prices = []
        
        for product in kit_products:
            name = product.get('product_name', '')
            if name:
                product_names.append(name)
            
            price = product.get('price_analysis', {}).get('average_price', 0)
            individual_prices.append(price)
        
        # Calcular preço total
        total_price = sum(individual_prices)
        
        # Aplicar desconto
        discount = 5
        kit_price = total_price * (1 - (discount / 100))
        
        # Calcular pontuação média
        avg_score = sum(p.get('overall_score', 0) for p in kit_products) / len(kit_products)
        
        # Criar kit
        kit = {
            'kit_name': f"Kit Premium {i+1}",
            'products': product_names,
            'individual_prices': individual_prices,
            'total_price': total_price,
            'kit_price': kit_price,
            'discount': discount,
            'average_score': avg_score,
            'reasoning': "Kit composto pelos produtos mais bem avaliados"
        }
        
        kits.append(kit)
    
    return kits

def generate_smart_kits(product_analyses, max_kits=5, kit_size=3, use_ai=True, model="gpt-3.5-turbo"):
    """
    Versão melhorada para gerar kits de produtos usando regras de negócio
    e inteligência artificial.
    
    Args:
        product_analyses: Lista de análises de produtos
        max_kits: Número máximo de kits a gerar
        kit_size: Tamanho de cada kit
        use_ai: Se True, usa IA para análise
        model: Modelo de IA a usar
        
    Returns:
        Lista de kits gerados
    """
    # Filtrar produtos válidos e bem avaliados
    valid_products = [p for p in product_analyses if p.get('found', True)]
    
    # Verificar se temos produtos suficientes
    if len(valid_products) < kit_size:
        log_warning(f"Produtos insuficientes para gerar kits ({len(valid_products)})")
        return []
    
    # Agrupar produtos por categoria/tipo
    categorized_products = categorize_products(valid_products)
    
    # Identificar os melhores produtos por categoria
    top_products_by_category = {}
    for category, products in categorized_products.items():
        # Ordena os produtos da categoria por pontuação
        sorted_products = sorted(products, key=lambda x: x.get('overall_score', 0), reverse=True)
        top_products_by_category[category] = sorted_products[:min(5, len(sorted_products))]
    
    # Gerar kits usando IA, se disponível
    if use_ai and openai.api_key:
        try:
            ai_kits = generate_kits_with_ai(valid_products, categorized_products, max_kits, kit_size, model)
            
            if ai_kits:
                # Enriquecer os kits gerados pela IA
                enriched_kits = enrich_kits(ai_kits, valid_products)
                if len(enriched_kits) >= max_kits // 2:  # Se tiver pelo menos metade dos kits solicitados
                    return enriched_kits
        except Exception as e:
            log_error(f"Erro ao gerar kits com IA", e)
    
    # Se a IA falhou ou não está disponível, usar método híbrido
    return generate_hybrid_kits(valid_products, categorized_products, top_products_by_category, max_kits, kit_size)

def categorize_products(products):
    """
    Categoriza produtos por tipo/categoria.
    
    Args:
        products: Lista de análises de produtos
        
    Returns:
        Dicionário de produtos agrupados por categoria
    """
    category_dict = {}
    
    for product in products:
        # Extrair nome do produto
        product_name = product.get('product_name', '').lower()
        
        # Detectar categoria
        found_category = False
        
        # Dicionário de categorias e palavras-chave
        categories = {
            'Eletrônicos': ['celular', 'smartphone', 'tv', 'televisão', 'monitor', 'tablet', 'notebook', 'laptop', 'fone', 'headphone'],
            'Informática': ['computador', 'pc', 'teclado', 'mouse', 'impressora', 'scanner', 'webcam', 'hd', 'ssd', 'pendrive'],
            'Móveis': ['mesa', 'cadeira', 'sofá', 'poltrona', 'armário', 'estante', 'cama', 'guarda-roupa', 'criado-mudo'],
            'Eletrodomésticos': ['geladeira', 'fogão', 'microondas', 'liquidificador', 'batedeira', 'cafeteira', 'aspirador'],
            'Ferramentas': ['martelo', 'chave', 'parafusadeira', 'furadeira', 'alicate', 'serra', 'esmerilhadeira'],
            'Decoração': ['tapete', 'cortina', 'quadro', 'luminária', 'espelho', 'vaso', 'almofada'],
            'Vestuário': ['camisa', 'camiseta', 'calça', 'vestido', 'bermuda', 'jaqueta', 'casaco', 'sapato', 'tênis'],
            'Brinquedos': ['boneca', 'carrinho', 'jogo', 'puzzle', 'quebra-cabeça', 'lego', 'nerf']
        }
        
        for category, keywords in categories.items():
            if any(keyword in product_name for keyword in keywords):
                # Adicionar ao dicionário de categorias
                if category not in category_dict:
                    category_dict[category] = []
                category_dict[category].append(product)
                found_category = True
                break
        
        # Se não encontrou categoria, adicionar a "Diversos"
        if not found_category:
            if 'Diversos' not in category_dict:
                category_dict['Diversos'] = []
            category_dict['Diversos'].append(product)
    
    return category_dict

def generate_kits_with_ai(products, categorized_products, max_kits, kit_size, model):
    """
    Gera kits usando IA com um prompt melhorado.
    
    Args:
        products: Lista completa de produtos
        categorized_products: Produtos categorizados
        max_kits: Número máximo de kits
        kit_size: Tamanho de cada kit
        model: Modelo de IA a usar
        
    Returns:
        Lista de kits gerados
    """
    # Ordenar produtos por pontuação
    sorted_products = sorted(
        products, 
        key=lambda x: x.get('overall_score', 0), 
        reverse=True
    )
    
    # Obter os top produtos
    top_products = sorted_products[:20]  # Aumentar para 20 para mais variedade
    
    # Preparar informações dos produtos para o prompt
    product_infos = []
    for i, product in enumerate(top_products):
        name = product.get('product_name', f"Produto {i+1}")
        price = product.get('price_analysis', {}).get('average_price', 0)
        score = product.get('overall_score', 0)
        sold = product.get('demand_analysis', {}).get('average_sold', 0)
        
        # Tentar identificar a categoria
        category = "Diversos"
        for cat_name, cat_products in categorized_products.items():
            if product in cat_products:
                category = cat_name
                break
        
        product_infos.append(f"ID: {i+1}, Nome: {name}, Categoria: {category}, Preço: R$ {price:.2f}, Score: {score:.1f}, Vendas: {sold:.0f}")
    
    products_text = "\n".join(product_infos)
    
    # Criar prompt inteligente para a IA
    prompt = f"""
    Como especialista em estratégia de vendas para o Mercado Livre, crie {max_kits} kits diferentes de produtos que maximizem as vendas e a rentabilidade.
    
    PRODUTOS DISPONÍVEIS:
    {products_text}
    
    REGRAS PARA CRIAÇÃO DE KITS:
    1. Cada kit deve conter exatamente {kit_size} produtos
    2. Os produtos em um kit devem ser complementares e fazer sentido juntos
    3. Priorize produtos com maior pontuação (score) e demanda (vendas)
    4. Crie variedade entre os kits (evite repetir os mesmos produtos)
    5. Considere diferentes perfis de clientes (iniciantes, avançados, corporativos, etc.)
    6. Produtos da mesma categoria geralmente combinam bem, mas kits cross-categoria também podem ser interessantes
    7. Aplique um desconto entre 5-15% sobre o valor total, sendo maior o desconto para kits de maior valor
    8. O nome do kit deve ser atrativo e comunicar valor para o cliente
    
    FORMATO DA RESPOSTA:
    Retorne um array JSON contendo {max_kits} kits, cada um no seguinte formato:
    
    [
        {{
            "kit_name": "Nome do Kit 1",
            "target_audience": "Público-alvo deste kit",
            "products": [IDs dos produtos, ex: [1, 5, 7]],
            "individual_prices": [preços individuais],
            "total_price": preço total sem desconto,
            "kit_price": preço com desconto,
            "discount": percentual de desconto,
            "average_score": pontuação média do kit,
            "reasoning": "Explicação do porquê esse kit funcionaria bem",
            "marketing_pitch": "Texto curto para anunciar este kit"
        }},
        // mais kits...
    ]
    
    Observe que os produtos em cada kit devem realmente combinar bem e fazer sentido para o consumidor.
    Retorne apenas o JSON, sem texto adicional.
    """
    
    # Chamar a API da OpenAI
    response = openai.ChatCompletion.create(
        model=model,
        messages=[
            {"role": "system", "content": "Você é um especialista em estratégia de vendas para e-commerce, especializado em criar kits de produtos que maximizam as vendas e rentabilidade."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.4,
        max_tokens=2000
    )
    
    # Extrair a resposta
    ai_response = response.choices[0].message['content']
    
    # Processar a resposta JSON
    kits = parse_kit_recommendations_enhanced(ai_response, top_products)
    
    if kits:
        log_success(f"Gerados {len(kits)} kits com IA")
        return kits
    else:
        log_warning("Falha ao gerar kits com IA, resposta inválida")
        return []

def parse_kit_recommendations_enhanced(response_text, products):
    """
    Versão melhorada para processar a resposta da IA.
    
    Args:
        response_text: Texto da resposta da IA
        products: Lista de produtos disponíveis
        
    Returns:
        Lista de kits recomendados
    """
    if not response_text:
        return []
    
    try:
        # Limpar o texto para obter apenas o JSON
        cleaned_text = response_text.strip()
        
        # Remover marcadores de código se presentes
        if cleaned_text.startswith('```json'):
            cleaned_text = cleaned_text[7:]
        elif cleaned_text.startswith('```'):
            cleaned_text = cleaned_text[3:]
            
        if cleaned_text.endswith('```'):
            cleaned_text = cleaned_text[:-3]
        
        # Tentar corrigir JSON malformado (problemas comuns)
        # Substituir aspas simples por aspas duplas
        cleaned_text = cleaned_text.replace("'", '"')
        
        # Colocar aspas em chaves numéricas que estejam sem aspas
        cleaned_text = re.sub(r'(\s*})(\s*,?\s*)([\d]+)(\s*):', r'\1\2"\3"\4:', cleaned_text)
        
        # Substituir NaN ou undefined por null
        cleaned_text = re.sub(r'\bNaN\b', 'null', cleaned_text)
        cleaned_text = re.sub(r'\bundefined\b', 'null', cleaned_text)
        
        # Analisar o JSON
        try:
            kits_data = json.loads(cleaned_text)
        except json.JSONDecodeError as e:
            debug_print(f"Erro ao decodificar JSON: {str(e)}")
            # Tentar extrair arrays JSON usando regex como último recurso
            pattern = r'\[\s*{(.+?)}\s*\]'
            match = re.search(pattern, cleaned_text, re.DOTALL)
            if match:
                try:
                    json_text = "[{" + match.group(1) + "}]"
                    kits_data = json.loads(json_text)
                except:
                    return []
            else:
                return []
        
        # Verificar se é uma lista
        if not isinstance(kits_data, list):
            debug_print("Resposta não é uma lista")
            return []
        
        # Processar cada kit
        processed_kits = []
        
        for kit_data in kits_data:
            try:
                # Extrair IDs dos produtos
                product_ids = kit_data.get('products', [])
                if not product_ids:
                    continue
                
                product_names = []
                
                # Converter IDs para nomes de produtos
                for pid in product_ids:
                    if isinstance(pid, int) and 1 <= pid <= len(products):
                        product_names.append(products[pid-1].get('product_name', f"Produto {pid}"))
                    else:
                        product_names.append(f"Produto {pid}")
                
                # Preços individuais
                individual_prices = kit_data.get('individual_prices', [])
                
                # Se não tiver preços ou o número for diferente, recalcular
                if not individual_prices or len(individual_prices) != len(product_ids):
                    individual_prices = []
                    for pid in product_ids:
                        if isinstance(pid, int) and 1 <= pid <= len(products):
                            price = products[pid-1].get('price_analysis', {}).get('average_price', 0)
                            individual_prices.append(price)
                        else:
                            individual_prices.append(0)
                
                # Calcular total e preço com desconto
                total_price = kit_data.get('total_price', sum(individual_prices))
                discount = kit_data.get('discount', 5)
                kit_price = kit_data.get('kit_price', total_price * (1 - discount/100))
                
                # Pontuação média
                scores = []
                for pid in product_ids:
                    if isinstance(pid, int) and 1 <= pid <= len(products):
                        scores.append(products[pid-1].get('overall_score', 0))
                
                avg_score = kit_data.get('average_score', sum(scores) / len(scores) if scores else 0)
                
                # Verificar e definir target_audience se não existir
                target_audience = kit_data.get('target_audience', "Clientes gerais")
                
                # Verificar e definir marketing_pitch se não existir
                marketing_pitch = kit_data.get('marketing_pitch', f"Kit completo com {len(product_names)} produtos essenciais com {discount}% de desconto!")
                
                # Criar kit processado
                processed_kit = {
                    'kit_name': kit_data.get('kit_name', f"Kit Premium {len(processed_kits) + 1}"),
                    'target_audience': target_audience,
                    'products': product_names,
                    'individual_prices': individual_prices,
                    'total_price': total_price,
                    'kit_price': kit_price,
                    'discount': discount,
                    'average_score': avg_score,
                    'reasoning': kit_data.get('reasoning', "Kit com produtos complementares"),
                    'marketing_pitch': marketing_pitch
                }
                
                processed_kits.append(processed_kit)
            except Exception as e:
                debug_print(f"Erro ao processar kit: {str(e)}")
        
        return processed_kits
        
    except Exception as e:
        debug_print(f"Erro ao processar kits da IA: {str(e)}")
        return []

def generate_hybrid_kits(products, categorized_products, top_by_category, max_kits, kit_size):
    """
    Gera kits usando uma abordagem híbrida inteligente sem IA.
    
    Args:
        products: Lista de produtos
        categorized_products: Produtos categorizados
        top_by_category: Melhores produtos por categoria
        max_kits: Número máximo de kits
        kit_size: Tamanho de cada kit
        
    Returns:
        Lista de kits gerados
    """
    kits = []
    kit_count = 0
    
    # 1. Criar kits por categoria (para as categorias com produtos suficientes)
    for category, cat_products in categorized_products.items():
        if len(cat_products) >= kit_size and kit_count < max_kits:
            # Ordenar produtos da categoria por pontuação
            sorted_cat_products = sorted(cat_products, key=lambda x: x.get('overall_score', 0), reverse=True)
            
            # Selecionar os melhores produtos da categoria
            kit_products = sorted_cat_products[:kit_size]
            
            kit = create_kit_from_products(
                kit_products,
                f"Kit {category} Premium",
                f"Clientes interessados em {category}",
                f"Kit completo com {kit_size} produtos essenciais de {category}",
                "Kit composto pelos melhores produtos da categoria"
            )
            
            kits.append(kit)
            kit_count += 1
    
    # 2. Criar kits complementares entre categorias relacionadas
    if kit_count < max_kits:
        category_pairs = [
            ('Eletrônicos', 'Informática'),
            ('Móveis', 'Decoração'),
            ('Ferramentas', 'Eletrodomésticos')
        ]
        
        for cat1, cat2 in category_pairs:
            if kit_count >= max_kits:
                break
                
            # Verificar se temos ambas as categorias
            if cat1 in top_by_category and cat2 in top_by_category:
                # Pegar os melhores produtos de cada categoria
                products1 = top_by_category[cat1]
                products2 = top_by_category[cat2]
                
                # Verificar se temos produtos suficientes
                if len(products1) >= 1 and len(products2) >= 1:
                    # Determinar quantos produtos de cada categoria
                    cat1_count = min(len(products1), kit_size // 2 + kit_size % 2)
                    cat2_count = min(len(products2), kit_size - cat1_count)
                    
                    # Ajustar cat1_count se necessário
                    if cat1_count + cat2_count < kit_size and len(products1) > cat1_count:
                        cat1_count = min(len(products1), kit_size - cat2_count)
                    
                    # Selecionar produtos
                    kit_products = products1[:cat1_count] + products2[:cat2_count]
                    
                    if len(kit_products) == kit_size:
                        kit = create_kit_from_products(
                            kit_products,
                            f"Kit {cat1} & {cat2}",
                            f"Clientes interessados em {cat1} e {cat2}",
                            f"Combinação perfeita de produtos de {cat1} e {cat2} com desconto especial",
                            f"Kit misto que combina o melhor de {cat1} e {cat2}"
                        )
                        
                        kits.append(kit)
                        kit_count += 1
    
    # 3. Criar kits de iniciante, intermediário e avançado
    if kit_count < max_kits and len(products) >= kit_size * 3:
        # Ordenar todos os produtos por pontuação
        sorted_all = sorted(products, key=lambda x: x.get('overall_score', 0), reverse=True)
        
        # Ordenar por preço para kits de diferentes níveis
        sorted_by_price = sorted(products, key=lambda x: x.get('price_analysis', {}).get('average_price', 0))
        
        # Verificar se temos produtos suficientes
        if len(sorted_by_price) >= kit_size * 3:
            # Kit Iniciante (produtos mais baratos)
            if kit_count < max_kits:
                kit_products = sorted_by_price[:kit_size]
                kit = create_kit_from_products(
                    kit_products,
                    "Kit Iniciante",
                    "Clientes iniciantes com orçamento limitado",
                    "Kit perfeito para quem está começando, com ótimo custo-benefício",
                    "Kit econômico com produtos essenciais para iniciantes"
                )
                kits.append(kit)
                kit_count += 1
            
            # Kit Intermediário (produtos de preço médio)
            if kit_count < max_kits:
                mid_start = len(sorted_by_price) // 3
                kit_products = sorted_by_price[mid_start:mid_start+kit_size]
                kit = create_kit_from_products(
                    kit_products,
                    "Kit Intermediário",
                    "Clientes com experiência média",
                    "Solução completa com ótimo equilíbrio entre custo e desempenho",
                    "Kit com produtos de qualidade intermediária, excelente custo-benefício"
                )
                kits.append(kit)
                kit_count += 1
            
            # Kit Avançado (produtos mais caros/premium)
            if kit_count < max_kits:
                kit_products = sorted_by_price[-kit_size:]
                kit = create_kit_from_products(
                    kit_products,
                    "Kit Premium",
                    "Clientes exigentes que buscam o melhor",
                    "O melhor conjunto de produtos premium com desconto exclusivo",
                    "Kit com produtos top de linha para quem busca excelência"
                )
                kits.append(kit)
                kit_count += 1
    
    # 4. Kit com os produtos mais vendidos
    if kit_count < max_kits:
        # Ordenar por número de vendas
        sorted_by_sales = sorted(
            products, 
            key=lambda x: x.get('demand_analysis', {}).get('average_sold', 0),
            reverse=True
        )
        
        if len(sorted_by_sales) >= kit_size:
            kit_products = sorted_by_sales[:kit_size]
            kit = create_kit_from_products(
                kit_products,
                "Kit Mais Vendidos",
                "Todos os clientes",
                "Os produtos mais vendidos e bem avaliados com desconto especial",
                "Kit com os produtos mais populares e bem-sucedidos do mercado"
            )
            kits.append(kit)
            kit_count += 1
    
    # 5. Preencher com kits aleatórios até atingir o máximo
    while kit_count < max_kits:
        # Tentar criar um kit aleatório
        if len(products) >= kit_size:
            # Selecionar produtos aleatoriamente
            random_products = random.sample(products, kit_size)
            
            kit = create_kit_from_products(
                random_products,
                f"Kit Especial {kit_count + 1}",
                "Clientes diversos",
                "Combinação especial de produtos com desconto exclusivo",
                "Kit variado com produtos de diferentes categorias"
            )
            
            kits.append(kit)
            kit_count += 1
        else:
            break
    
    return kits

def create_kit_from_products(products, name, audience, pitch, reasoning):
    """
    Cria um kit a partir de uma lista de produtos.
    
    Args:
        products: Lista de produtos
        name: Nome do kit
        audience: Público-alvo
        pitch: Texto de marketing
        reasoning: Justificativa
        
    Returns:
        Dicionário com o kit
    """
    # Extrair nomes e preços
    product_names = []
    individual_prices = []
    
    for product in products:
        product_name = product.get('product_name', '')
        if product_name:
            product_names.append(product_name)
        
        price = product.get('price_analysis', {}).get('average_price', 0)
        individual_prices.append(price)
    
    # Calcular preço total
    total_price = sum(individual_prices)
    
    # Calcular desconto baseado no preço total
    if total_price > 1000:
        discount = 10
    elif total_price > 500:
        discount = 8
    else:
        discount = 5
    
    # Aplicar desconto
    kit_price = total_price * (1 - (discount / 100))
    
    # Calcular pontuação média
    avg_score = sum(p.get('overall_score', 0) for p in products) / len(products) if products else 0
    
    # Criar kit
    kit = {
        'kit_name': name,
        'target_audience': audience,
        'products': product_names,
        'individual_prices': individual_prices,
        'total_price': total_price,
        'kit_price': kit_price,
        'discount': discount,
        'average_score': avg_score,
        'reasoning': reasoning,
        'marketing_pitch': pitch
    }
    
    return kit

def enrich_kits(kits, products):
    """
    Enriquece kits com informações adicionais.
    
    Args:
        kits: Lista de kits
        products: Lista de produtos disponíveis
        
    Returns:
        Lista de kits enriquecidos
    """
    enriched_kits = []
    
    for kit in kits:
        # Garantir que temos todos os campos necessários
        if 'target_audience' not in kit:
            kit['target_audience'] = "Clientes em geral"
        
        if 'marketing_pitch' not in kit:
            product_count = len(kit.get('products', []))
            discount = kit.get('discount', 5)
            kit['marketing_pitch'] = f"Kit completo com {product_count} produtos essenciais com {discount}% de desconto!"
        
        # Enriquecer o nome do kit se for muito simples
        if len(kit.get('kit_name', '')) < 10:
            product_types = []
            for product_name in kit.get('products', []):
                product_type = classify_product_type(product_name)
                if product_type != "Diversos" and product_type not in product_types:
                    product_types.append(product_type)
            
            if product_types:
                kit['kit_name'] = f"Kit {' & '.join(product_types[:2])} {kit['kit_name']}"
        
        # Ajustar o desconto para ser múltiplo de 5 para marketing
        if 'discount' in kit:
            kit['discount'] = round(kit['discount'] / 5) * 5
            if kit['discount'] < 5:
                kit['discount'] = 5
            elif kit['discount'] > 15:
                kit['discount'] = 15
            
            # Recalcular o preço do kit
            kit['kit_price'] = kit.get('total_price', 0) * (1 - kit['discount']/100)
        
        enriched_kits.append(kit)
    
    return enriched_kits

# ------------------------------------------------------
# EXPORTAÇÃO PARA EXCEL
# ------------------------------------------------------

def export_analysis_to_excel(analyses, output_path):
    """
    Exporta análises de produtos para Excel.
    
    Args:
        analyses: Lista de análises de produtos
        output_path: Caminho para salvar o arquivo
        
    Returns:
        True se exportou com sucesso, False caso contrário
    """
    try:
        # Verificar se temos dados para exportar
        if not analyses:
            log_warning("Nenhuma análise para exportar")
            return False
        
        # Preparar dados para o DataFrame
        data = []
        
        for analysis in analyses:
            if analysis.get('found', False):
                row = {
                    'Produto': analysis.get('product_name', ''),
                    'Preço Inicial (PDF)': analysis.get('initial_price', 'N/A'),
                    'Preço Médio (ML)': safe_get(analysis, ['price_analysis', 'average_price'], 0),
                    'Margem (%)': safe_get(analysis, ['price_analysis', 'average_margin'], 0),
                    'Análise de Preço': safe_get(analysis, ['price_analysis', 'details'], 'N/A'),
                    'Vendedores de Alto Nível (%)': safe_get(analysis, ['competition_analysis', 'high_level_sellers'], 0),
                    'Análise de Concorrência': safe_get(analysis, ['competition_analysis', 'details'], 'N/A'),
                    'Vendas Médias': safe_get(analysis, ['demand_analysis', 'average_sold'], 0),
                    'Análise de Demanda': safe_get(analysis, ['demand_analysis', 'details'], 'N/A'),
                    'Pontuação Geral': analysis.get('overall_score', 0),
                    'Recomendação': analysis.get('recommendation', 'N/A')
                }
            else:
                row = {
                    'Produto': analysis.get('product_name', ''),
                    'Preço Inicial (PDF)': analysis.get('initial_price', 'N/A'),
                    'Preço Médio (ML)': 0,
                    'Margem (%)': 0,
                    'Análise de Preço': 'N/A',
                    'Vendedores de Alto Nível (%)': 0,
                    'Análise de Concorrência': 'N/A',
                    'Vendas Médias': 0,
                    'Análise de Demanda': 'N/A',
                    'Pontuação Geral': 0,
                    'Recomendação': analysis.get('recommendation', 'Não encontrado')
                }
            
            data.append(row)
        
        # Criar DataFrame
        df = pd.DataFrame(data)
        
        # Criar arquivo Excel
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, sheet_name='Análise de Produtos', index=False)
        
        # Formatar planilha
        workbook = writer.book
        worksheet = writer.sheets['Análise de Produtos']
        
        # Ajustar largura das colunas
        for i, column in enumerate(df.columns):
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            ) + 2
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = max_length
        
        # Adicionar cores com base nas pontuações
        try:
            score_col_idx = df.columns.get_loc('Pontuação Geral') + 1
            
            for row_idx in range(2, len(df) + 2):
                score_cell = worksheet.cell(row=row_idx, column=score_col_idx)
                try:
                    score = float(score_cell.value) if score_cell.value else 0
                except (ValueError, TypeError):
                    score = 0
                
                # Escolher cor com base na pontuação
                if score >= 7:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
                elif score >= 5:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo
                elif score >= 3:
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Laranja
                else:
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Vermelho
                
                # Aplicar cor à linha inteira
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill
        except Exception as e:
            debug_print(f"Erro ao colorir células: {str(e)}")
        
        # Formatar cabeçalhos
        try:
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
        except Exception as e:
            debug_print(f"Erro ao formatar cabeçalhos: {str(e)}")
        
        # Salvar o arquivo
        writer.close()
        
        log_success(f"Análises exportadas para: {output_path}")
        return True
        
    except Exception as e:
        log_error(f"Erro ao exportar análises para Excel", e)
        try:
            # Tentar salvar como CSV (fallback)
            csv_path = output_path.replace('.xlsx', '.csv')
            pd.DataFrame(data).to_csv(csv_path, index=False)
            log_success(f"Análises exportadas para CSV: {csv_path}")
            return True
        except Exception as csv_error:
            log_error(f"Erro ao exportar para CSV", csv_error)
            return False

def export_analysis_to_excel_enhanced(analyses, output_path):
    """
    Versão aprimorada da exportação de análises de produtos para Excel,
    com formatação melhorada e visualizações.
    
    Args:
        analyses: Lista de análises de produtos
        output_path: Caminho para salvar o arquivo
        
    Returns:
        True se exportou com sucesso, False caso contrário
    """
    try:
        # Verificar se temos dados para exportar
        if not analyses:
            log_warning("Nenhuma análise para exportar")
            return False
        
        # Preparar dados para o DataFrame
        data = []
        
        for analysis in analyses:
            if analysis.get('found', False):
                row = {
                    'Produto': analysis.get('product_name', ''),
                    'Preço Inicial (PDF)': analysis.get('initial_price', 'N/A'),
                    'Preço Médio (ML)': safe_get(analysis, ['price_analysis', 'average_price'], 0),
                    'Margem (%)': safe_get(analysis, ['price_analysis', 'average_margin'], 0),
                    'Análise de Preço': safe_get(analysis, ['price_analysis', 'details'], 'N/A'),
                    'Pontuação Preço': safe_get(analysis, ['price_analysis', 'score'], 0),
                    'Vendedores de Alto Nível (%)': safe_get(analysis, ['competition_analysis', 'high_level_sellers'], 0),
                    'Análise de Concorrência': safe_get(analysis, ['competition_analysis', 'details'], 'N/A'),
                    'Pontuação Concorrência': safe_get(analysis, ['competition_analysis', 'score'], 0),
                    'Vendas Médias': safe_get(analysis, ['demand_analysis', 'average_sold'], 0),
                    'Análise de Demanda': safe_get(analysis, ['demand_analysis', 'details'], 'N/A'),
                    'Pontuação Demanda': safe_get(analysis, ['demand_analysis', 'score'], 0),
                    'Pontuação Geral': analysis.get('overall_score', 0),
                    'Recomendação': analysis.get('recommendation', 'N/A'),
                    'Sugestões': format_suggestions(safe_get(analysis, ['improvement_suggestions'], []))
                }
            else:
                row = {
                    'Produto': analysis.get('product_name', ''),
                    'Preço Inicial (PDF)': analysis.get('initial_price', 'N/A'),
                    'Preço Médio (ML)': 0,
                    'Margem (%)': 0,
                    'Análise de Preço': 'N/A',
                    'Pontuação Preço': 0,
                    'Vendedores de Alto Nível (%)': 0,
                    'Análise de Concorrência': 'N/A',
                    'Pontuação Concorrência': 0,
                    'Vendas Médias': 0,
                    'Análise de Demanda': 'N/A',
                    'Pontuação Demanda': 0,
                    'Pontuação Geral': 0,
                    'Recomendação': analysis.get('recommendation', 'Não encontrado'),
                    'Sugestões': 'N/A'
                }
            
            data.append(row)
        
        # Criar DataFrame
        df = pd.DataFrame(data)
        
        # Adicionar uma aba de resumo
        summary_data = create_analysis_summary(analyses)
        
        # Salvar no Excel com múltiplas abas
        save_enhanced_excel(df, summary_data, output_path)
        
        log_success(f"Análises exportadas para: {output_path}")
        return True
        
    except Exception as e:
        log_error(f"Erro ao exportar análises para Excel", e)
        try:
            # Tentar salvar como CSV (fallback)
            csv_path = output_path.replace('.xlsx', '.csv')
            pd.DataFrame(data).to_csv(csv_path, index=False)
            log_success(f"Análises exportadas para CSV: {csv_path}")
            return True
        except Exception as csv_error:
            log_error(f"Erro ao exportar para CSV", csv_error)
            return False

def format_suggestions(suggestions):
    """
    Formata a lista de sugestões para exibição no Excel.
    
    Args:
        suggestions: Lista de sugestões
        
    Returns:
        String formatada
    """
    if not suggestions:
        return "N/A"
    
    if isinstance(suggestions, list):
        return "\n• " + "\n• ".join(suggestions)
    else:
        return str(suggestions)

def create_analysis_summary(analyses):
    """
    Cria um resumo das análises para a aba de resumo.
    
    Args:
        analyses: Lista de análises de produtos
        
    Returns:
        DataFrame com o resumo
    """
    # Filtrar apenas produtos encontrados
    valid_analyses = [a for a in analyses if a.get('found', False)]
    
    if not valid_analyses:
        return pd.DataFrame()
    
    # Estatísticas gerais
    total_products = len(analyses)
    found_products = len(valid_analyses)
    not_found = total_products - found_products
    
    # Contagem por recomendação
    rec_counts = {}
    for a in valid_analyses:
        rec = a.get('recommendation', 'N/A')
        rec_counts[rec] = rec_counts.get(rec, 0) + 1
    
    # Médias
    avg_price = sum(safe_get(a, ['price_analysis', 'average_price'], 0) for a in valid_analyses) / len(valid_analyses) if valid_analyses else 0
    avg_margin = sum(safe_get(a, ['price_analysis', 'average_margin'], 0) for a in valid_analyses) / len(valid_analyses) if valid_analyses else 0
    avg_score = sum(a.get('overall_score', 0) for a in valid_analyses) / len(valid_analyses) if valid_analyses else 0
    
    # Produtos com maior pontuação
    top_products = sorted(valid_analyses, key=lambda x: x.get('overall_score', 0), reverse=True)[:10]
    top_products_data = []
    
    for i, product in enumerate(top_products):
        top_products_data.append({
            'Ranking': i + 1,
            'Produto': product.get('product_name', ''),
            'Pontuação': product.get('overall_score', 0),
            'Preço Médio': safe_get(product, ['price_analysis', 'average_price'], 0),
            'Margem': safe_get(product, ['price_analysis', 'average_margin'], 0),
            'Vendas': safe_get(product, ['demand_analysis', 'average_sold'], 0),
            'Recomendação': product.get('recommendation', 'N/A')
        })
    
    # Produtos com maior demanda
    top_demand_products = sorted(valid_analyses, key=lambda x: safe_get(x, ['demand_analysis', 'average_sold'], 0), reverse=True)[:10]
    top_demand_data = []
    
    for i, product in enumerate(top_demand_products):
        top_demand_data.append({
            'Ranking': i + 1,
            'Produto': product.get('product_name', ''),
            'Vendas': safe_get(product, ['demand_analysis', 'average_sold'], 0),
            'Pontuação': product.get('overall_score', 0),
            'Preço Médio': safe_get(product, ['price_analysis', 'average_price'], 0),
            'Recomendação': product.get('recommendation', 'N/A')
        })
    
    # Produtos com melhor margem
    top_margin_products = sorted(valid_analyses, key=lambda x: safe_get(x, ['price_analysis', 'average_margin'], 0), reverse=True)[:10]
    top_margin_data = []
    
    for i, product in enumerate(top_margin_products):
        top_margin_data.append({
            'Ranking': i + 1,
            'Produto': product.get('product_name', ''),
            'Margem': safe_get(product, ['price_analysis', 'average_margin'], 0),
            'Pontuação': product.get('overall_score', 0),
            'Preço Médio': safe_get(product, ['price_analysis', 'average_price'], 0),
            'Recomendação': product.get('recommendation', 'N/A')
        })
    
    # Estatísticas gerais em DataFrame
    general_stats = [{
        'Métrica': 'Total de Produtos Analisados',
        'Valor': total_products
    }, {
        'Métrica': 'Produtos Encontrados no ML',
        'Valor': found_products
    }, {
        'Métrica': 'Produtos Não Encontrados',
        'Valor': not_found
    }, {
        'Métrica': 'Preço Médio (R$)',
        'Valor': avg_price
    }, {
        'Métrica': 'Margem Média (%)',
        'Valor': avg_margin
    }, {
        'Métrica': 'Pontuação Média',
        'Valor': avg_score
    }]
    
    # Adicionar contagens por recomendação
    for rec, count in rec_counts.items():
        general_stats.append({
            'Métrica': f'Produtos {rec}',
            'Valor': count
        })
    
    return {
        'general_stats': pd.DataFrame(general_stats),
        'top_products': pd.DataFrame(top_products_data),
        'top_demand': pd.DataFrame(top_demand_data),
        'top_margin': pd.DataFrame(top_margin_data)
    }

def save_enhanced_excel(df, summary_data, output_path):
    """
    Salva os dados em um arquivo Excel com formatação melhorada.
    
    Args:
        df: DataFrame com os dados de produtos
        summary_data: Dados do resumo
        output_path: Caminho para salvar o arquivo
        
    Returns:
        None
    """
    try:
        # Criar arquivo Excel
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # Aba de resumo
        if summary_data and 'general_stats' in summary_data:
            # Estatísticas gerais
            summary_data['general_stats'].to_excel(writer, sheet_name='Resumo', index=False, startrow=1, startcol=1)
            
            # Top produtos por pontuação
            if 'top_products' in summary_data:
                summary_data['top_products'].to_excel(writer, sheet_name='Resumo', index=False, startrow=len(summary_data['general_stats']) + 5, startcol=1)
            
            # Top produtos por demanda
            if 'top_demand' in summary_data:
                summary_data['top_demand'].to_excel(writer, sheet_name='Resumo', index=False, startrow=1, startcol=6)
            
            # Top produtos por margem
            if 'top_margin' in summary_data:
                summary_data['top_margin'].to_excel(writer, sheet_name='Resumo', index=False, startrow=len(summary_data['general_stats']) + 5, startcol=6)
        
        # Aba de análise detalhada
        df.to_excel(writer, sheet_name='Análise de Produtos', index=False)
        
        # Obter o livro e as planilhas
        workbook = writer.book
        
        # Formatar a aba de resumo
        if summary_data and 'general_stats' in summary_data:
            format_summary_sheet(workbook, 'Resumo', summary_data)
        
        # Formatar a aba de análise
        format_analysis_sheet(workbook, 'Análise de Produtos', df)
        
        # Salvar o arquivo
        writer.close()
    
    except Exception as e:
        log_error(f"Erro ao salvar Excel aprimorado: {str(e)}")
        # Salvar de forma simples como fallback
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Análise de Produtos', index=False)

def format_summary_sheet(workbook, sheet_name, summary_data):
    """
    Formata a aba de resumo.
    
    Args:
        workbook: Livro do Excel
        sheet_name: Nome da aba
        summary_data: Dados do resumo
        
    Returns:
        None
    """
    try:
        worksheet = workbook[sheet_name]
        
        # Adicionar título
        worksheet['B1'] = "RESUMO DA ANÁLISE"
        title_cell = worksheet['B1']
        title_cell.font = Font(bold=True, size=16, color="000080")
        
        # Títulos das seções
        section_titles = {
            'B1': "RESUMO DA ANÁLISE",
            'B3': "ESTATÍSTICAS GERAIS",
            f"B{len(summary_data['general_stats']) + 5}": "TOP 10 PRODUTOS POR PONTUAÇÃO",
            'G3': "TOP 10 PRODUTOS POR DEMANDA",
            f"G{len(summary_data['general_stats']) + 5}": "TOP 10 PRODUTOS POR MARGEM"
        }
        
        for cell, title in section_titles.items():
            worksheet[cell] = title
            title_cell = worksheet[cell]
            title_cell.font = Font(bold=True, size=12, color="000080")
        
        # Formatar a tabela de estatísticas gerais
        stats_start_row = 4
        format_table_range(
            worksheet, 
            f"B{stats_start_row}:C{stats_start_row + len(summary_data['general_stats']) - 1}",
            "4F81BD", "FFFFFF"
        )
        
        # Formatar tabelas top produtos
        top_start_row = len(summary_data['general_stats']) + 6
        format_table_range(
            worksheet, 
            f"B{top_start_row}:H{top_start_row + 10}",
            "4F81BD", "FFFFFF"
        )
        
        format_table_range(
            worksheet, 
            f"G4:L13",
            "4F81BD", "FFFFFF"
        )
        
        format_table_range(
            worksheet, 
            f"G{top_start_row}:L{top_start_row + 10}",
            "4F81BD", "FFFFFF"
        )
        
        # Ajustar largura das colunas
        for col in range(1, 15):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 15
    
    except Exception as e:
        debug_print(f"Erro ao formatar aba de resumo: {str(e)}")

def format_analysis_sheet(workbook, sheet_name, df):
    """
    Formata a aba de análise de produtos.
    
    Args:
        workbook: Livro do Excel
        sheet_name: Nome da aba
        df: DataFrame com os dados
        
    Returns:
        None
    """
    try:
        worksheet = workbook[sheet_name]
        
        # Ajustar largura das colunas
        for i, column in enumerate(df.columns, 1):
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            ) + 2
            col_letter = get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)
        
        # Formatar cabeçalhos
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Formatar células com base nas pontuações
        score_columns = {
            'Pontuação Preço': 6,
            'Pontuação Concorrência': 9,
            'Pontuação Demanda': 12,
            'Pontuação Geral': 13
        }
        
        for row_idx in range(2, len(df) + 2):
            # Verificar se o produto foi encontrado
            found = worksheet.cell(row=row_idx, column=3).value != 0  # Preço médio > 0
            
            if not found:
                # Aplicar estilo de produto não encontrado
                gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = gray_fill
            else:
                # Aplicar estilo com base na pontuação geral
                score_cell = worksheet.cell(row=row_idx, column=score_columns['Pontuação Geral'])
                try:
                    score = float(score_cell.value) if score_cell.value else 0
                except (ValueError, TypeError):
                    score = 0
                
                # Escolher cor com base na pontuação
                if score >= 7:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
                elif score >= 5:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo
                elif score >= 3:
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Laranja
                else:
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Vermelho
                
                # Aplicar cor às células chave
                for col_name, col_idx in score_columns.items():
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill
                
                # Colorir a célula de recomendação
                recommendation_col = 14  # Coluna da recomendação
                rec_cell = worksheet.cell(row=row_idx, column=recommendation_col)
                rec_cell.fill = fill
                
                # Formatar células de pontuação com barra visual
                for col_name, col_idx in score_columns.items():
                    score_value = worksheet.cell(row=row_idx, column=col_idx).value
                    try:
                        score_num = float(score_value) if score_value else 0
                    except (ValueError, TypeError):
                        score_num = 0
                    
                    # Adicionar barras de pontuação (usando caracteres Unicode)
                    bar_count = int(score_num)
                    worksheet.cell(row=row_idx, column=col_idx).value = f"{score_num:.1f} {'■' * bar_count}{'□' * (10-bar_count)}"
        
        # Formatar colunas numéricas
        numeric_columns = {
            'Preço Inicial (PDF)': 2,
            'Preço Médio (ML)': 3,
            'Margem (%)': 4,
            'Vendedores de Alto Nível (%)': 7,
            'Vendas Médias': 10
        }
        
        for col_name, col_idx in numeric_columns.items():
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    value = float(cell.value) if cell.value not in ('N/A', None) else 0
                    
                    if 'Preço' in col_name:
                        cell.value = value
                        cell.number_format = 'R$ #,##0.00'
                    elif '%' in col_name:
                        cell.value = value
                        cell.number_format = '0.0%'
                    else:
                        cell.value = value
                        cell.number_format = '#,##0.0'
                except (ValueError, TypeError):
                    pass
        
        # Formatar colunas de texto longos para quebrar linhas
        text_columns = {
            'Análise de Preço': 5,
            'Análise de Concorrência': 8,
            'Análise de Demanda': 11,
            'Sugestões': 15
        }
        
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for col_name, col_idx in text_columns.items():
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = wrap_alignment
                
                # Aumentar altura da linha se necessário
                row_height = 15
                text = str(cell.value)
                if text and text != 'N/A':
                    lines = len(text.split('\n'))
                    if lines > 1:
                        row_height = max(row_height, lines * 15)
                
                worksheet.row_dimensions[row_idx].height = row_height
    
    except Exception as e:
        debug_print(f"Erro ao formatar aba de análise: {str(e)}")

def format_table_range(worksheet, range_str, header_color, header_text_color):
    """
    Formata uma tabela.
    
    Args:
        worksheet: Planilha do Excel
        range_str: String com o range da tabela
        header_color: Cor do cabeçalho
        header_text_color: Cor do texto do cabeçalho
        
    Returns:
        None
    """
    try:
        # Aplicar estilos alternados para as linhas
        start_col, start_row, end_col, end_row = range_reference_to_indices(range_str)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # Primeira linha (cabeçalho)
                if row == start_row:
                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                    cell.font = Font(bold=True, color=header_text_color)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Linhas de dados
                else:
                    # Estilo alternado
                    if (row - start_row) % 2 == 0:
                        cell.fill = PatternFill(start_color="E9EDF1", end_color="E9EDF1", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="D0D8E8", end_color="D0D8E8", fill_type="solid")
    
    except Exception as e:
        debug_print(f"Erro ao formatar tabela: {str(e)}")

def range_reference_to_indices(range_reference):
    """
    Converte uma referência de intervalo (ex: 'A1:C5') em índices de linha e coluna.
    
    Args:
        range_reference: Referência de intervalo no formato Excel
        
    Returns:
        Tupla com (start_col, start_row, end_col, end_row)
    """
    start, end = range_reference.split(':')
    
    # Extrair letras e números
    start_col_letters = ''.join(filter(str.isalpha, start))
    start_row = int(''.join(filter(str.isdigit, start)))
    
    end_col_letters = ''.join(filter(str.isalpha, end))
    end_row = int(''.join(filter(str.isdigit, end)))
    
    # Converter letras em números de coluna
    start_col = column_letter_to_index(start_col_letters)
    end_col = column_letter_to_index(end_col_letters)
    
    return start_col, start_row, end_col, end_row

def column_letter_to_index(column_letter):
    """
    Converte letra de coluna para índice.
    
    Args:
        column_letter: Letra da coluna (A, B, C, ..., AA, AB, etc.)
        
    Returns:
        Índice da coluna (1-based)
    """
    result = 0
    for i, char in enumerate(reversed(column_letter)):
        result += (ord(char) - 64) * (26 ** i)
    return result

def export_kits_to_excel(kits, output_path):
    """
    Exporta kits para Excel.
    
    Args:
        kits: Lista de kits
        output_path: Caminho para salvar o arquivo
        
    Returns:
        True se exportou com sucesso, False caso contrário
    """
    try:
        # Verificar se temos dados para exportar
        if not kits:
            log_warning("Nenhum kit para exportar")
            return False
        
        # Preparar dados para o DataFrame
        data = []
        
        for kit in kits:
            row = {
                'Nome do Kit': kit.get('kit_name', ''),
                'Produtos': format_product_list(kit.get('products', [])),
                'Preço Total Individual': kit.get('total_price', 0),
                'Preço do Kit': kit.get('kit_price', 0),
                'Desconto (%)': kit.get('discount', 0),
                'Pontuação Média': kit.get('average_score', 0),
                'Justificativa': kit.get('reasoning', 'Kit com produtos complementares')
            }
            
            data.append(row)
        
        # Criar DataFrame
        df = pd.DataFrame(data)
        
        # Criar arquivo Excel
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, sheet_name='Recomendações de Kits', index=False)
        
        # Formatar planilha
        workbook = writer.book
        worksheet = writer.sheets['Recomendações de Kits']
        
        # Ajustar largura das colunas
        for i, column in enumerate(df.columns):
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            ) + 2
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = max_length
        
        # Adicionar cores com base nas pontuações
        try:
            score_col_idx = df.columns.get_loc('Pontuação Média') + 1
            
            for row_idx in range(2, len(df) + 2):
                score_cell = worksheet.cell(row=row_idx, column=score_col_idx)
                try:
                    score = float(score_cell.value) if score_cell.value else 0
                except (ValueError, TypeError):
                    score = 0
                
                # Escolher cor com base na pontuação
                if score >= 7:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
                elif score >= 5:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo
                else:
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Vermelho
                
                # Aplicar cor à linha inteira
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill
        except Exception as e:
            debug_print(f"Erro ao colorir células: {str(e)}")
        
        # Formatar cabeçalhos
        try:
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
        except Exception as e:
            debug_print(f"Erro ao formatar cabeçalhos: {str(e)}")
        
        # Salvar o arquivo
        writer.close()
        
        log_success(f"Kits exportados para: {output_path}")
        return True
        
    except Exception as e:
        log_error(f"Erro ao exportar kits para Excel", e)
        try:
            # Tentar salvar como CSV (fallback)
            csv_path = output_path.replace('.xlsx', '.csv')
            pd.DataFrame(data).to_csv(csv_path, index=False)
            log_success(f"Kits exportados para CSV: {csv_path}")
            return True
        except Exception as csv_error:
            log_error(f"Erro ao exportar para CSV", csv_error)
            return False

def format_product_list(products):
    """
    Formata uma lista de produtos para melhor visualização no Excel.
    
    Args:
        products: Lista de produtos
        
    Returns:
        String formatada
    """
    if not products:
        return ""
    
    return "\n• " + "\n• ".join(products)

def export_kits_to_excel_enhanced(kits, output_path):
    """
    Versão melhorada da exportação de kits para Excel.
    
    Args:
        kits: Lista de kits
        output_path: Caminho para salvar o arquivo
        
    Returns:
        True se exportou com sucesso, False caso contrário
    """
    try:
        # Verificar se temos dados para exportar
        if not kits:
            log_warning("Nenhum kit para exportar")
            return False
        
        #
